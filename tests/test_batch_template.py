from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.batch_template import RULES_SHEET, TEMPLATE_ELEMENTS, export_batch_result, generate_template_workbook, parse_template_workbook, run_batch_optimization
from app.core import alloy_coeff, effective_bounds
from app.main import app


def workbook_bytes(
    *,
    duplicate_target: bool = False,
    merge_tasks: bool = False,
    formula_price: bool = False,
    task_header: list[str] | None = None,
    task_rows: list[list] | None = None,
    target_header: list[str] | None = None,
    target_rows: list[list] | None = None,
    endpoint_header: list[str] | None = None,
    endpoint_rows: list[list] | None = None,
    alloy_header: list[str] | None = None,
    alloy_rows: list[list] | None = None,
    price_rows: list[list] | None = None,
    rule_rows: list[list] | None = None,
) -> bytes:
    """Build a minimal user-facing upload template in memory."""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tasks = wb.create_sheet("01_批量任务")
    tasks.append(task_header or ["任务编号", "适用牌号", "厚度mm", "炉重t", "价格方案", "炼钢牌号", "备注"])
    for row in task_rows or [
        ["T001", "Q235B", 10, 150, "2026-05", "Q235B-1", "正常样例"],
        ["T002", "Q355C", 8, 150, "2026-05", "Q355C-1", "低合金样例"],
    ]:
        tasks.append(row)
    if merge_tasks:
        tasks.merge_cells("A2:B2")

    targets = wb.create_sheet("02_目标成分上下限")
    default_target_header = [
        "适用牌号",
        "最小厚度mm",
        "最大厚度mm",
        "炼钢牌号",
        "C下限",
        "C上限",
        "Si下限",
        "Si上限",
        "Mn下限",
        "Mn上限",
        "P上限",
        "S上限",
    ]
    targets.append(target_header or default_target_header)
    for row in target_rows or [
        ["Q235B", 1.5, 12, "Q235B-1", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
        ["Q355C", 3, 12, "Q355C-1", None, 0.16, 0.10, 0.12, 0.90, 0.92, 0.018, 0.010],
    ]:
        targets.append(row)
    if duplicate_target:
        targets.append(["Q235B-重复", 99, 100, "Q235B-1", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020])

    endpoints = wb.create_sheet("03_转炉终点与回收率")
    default_endpoint_header = [
            "适用牌号",
            "最小厚度mm",
            "最大厚度mm",
            "炼钢牌号",
            "C终点",
            "Mn终点",
            "Cr终点",
            "C回收率",
            "P回收率",
            "S回收率",
            "Si回收率",
            "Mn回收率",
            "V回收率",
            "Nb回收率",
            "Ti回收率",
            "Als回收率",
            "Alt回收率",
            "Ca回收率",
            "Cr回收率",
            "Ni回收率",
            "Cu回收率",
            "Mo回收率",
            "B回收率",
            "Sb回收率",
    ]
    endpoints.append(endpoint_header or default_endpoint_header)
    for row in endpoint_rows or [
        ["Q235B", 1.5, 12, "Q235B-1", 0.07, 0.12, 0, 0.90, 1.0, 1.0, 0.75, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
        ["Q355C", 3, 12, "Q355C-1", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
    ]:
        endpoints.append(row)

    alloys = wb.create_sheet("04_合金成分库")
    default_alloy_header = ["合金名称", "价格物料名", "启用", "投料方式", "袋重kg", "最大投加kg每t", "C", "Si", "Mn", "Cr", "P", "S", "N", "备注"]
    alloys.append(alloy_header or default_alloy_header)
    for alloy_row in alloy_rows or [
        ["硅锰", "硅锰", "是", "连续", 0, 30, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, "N 不参与优化"],
        ["硅铁", "硅铁", "是", "连续", 0, 20, 0.2, 72.23, None, None, 0.03, 0.02, None, ""],
        ["低碳锰铁", "低碳锰铁", "是", "整袋", 25, 25, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
    ]:
        alloys.append(alloy_row)

    prices = wb.create_sheet("05_价格表")
    prices.append(["价格方案", "物料名称", "价格日期", "价格元每吨"])
    for row in price_rows or [
        ["2026-05", "硅锰", "2026-05-07", 6130],
        ["2026-05", "硅铁", "2026-05-07", 5810],
        ["2026-05", "低碳锰铁", "2026-05-07", 10150],
    ]:
        prices.append(row)
    if formula_price:
        prices["D2"] = "=6000+130"

    rules = wb.create_sheet("06_填写说明与校验规则")
    rules.append(["模板版本", "1"])
    rules.append(["说明", "N 元素不参与合金优化。"])
    if rule_rows is not None:
        rule_sheet = wb.create_sheet(RULES_SHEET)
        rule_sheet.append(["规则项", "参数键", "值", "说明"])
        for row in rule_rows:
            rule_sheet.append(row)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_parse_template_workbook_builds_prevalidated_payload():
    report = parse_template_workbook(workbook_bytes())

    assert report["status"] == "ok"
    assert report["errors"] == []
    assert report["preview"]["taskCount"] == 2
    assert report["preview"]["alloyCount"] == 3
    assert "N" not in report["parsed"]["tasks"][0]["config"]["target"]
    assert report["parsed"]["tasks"][0]["config"]["residual"]["Si"] == 0


def test_parse_template_workbook_preserves_manual_aluminum_as_reference_record():
    report = parse_template_workbook(
        workbook_bytes(
            task_header=["任务编号", "适用牌号", "厚度mm", "炉重t", "价格方案", "炼钢牌号", "手工铝块kg/t", "备注"],
            task_rows=[
                ["T001", "Q235B", 10, 150, "2026-05", "Q235B-1", 1.25, "源表 AH"],
                ["T002", "Q355C", 8, 150, "2026-05", "Q355C-1", None, "未录入时按 0"],
            ],
            alloy_header=["合金名称", "价格物料名", "启用", "投料方式", "袋重kg", "最大投加kg每t", *TEMPLATE_ELEMENTS, "备注"],
            alloy_rows=[
                ["硅锰", "硅锰", "是", "连续", 0, 30, 1.72, 17.69, 65.66, 0.15, 0.02, None, None, None, None, None, None, None, None, None, None, None, None, ""],
                ["硅铁", "硅铁", "是", "连续", 0, 20, 0.2, 72.23, None, 0.03, 0.02, None, None, None, None, None, None, None, None, None, None, None, None, ""],
                ["低碳锰铁", "低碳锰铁", "是", "整袋", 25, 25, 0.64, None, 81.19, 0.20, 0.02, None, None, None, None, None, None, None, None, None, None, None, None, ""],
                ["铝块", "铝块", "是", "连续", 0, 10, None, None, None, 0, 0, None, None, 99, 99, None, None, None, None, None, None, None, None, "手工铝块单独记录"],
            ],
            price_rows=[
                ["2026-05", "硅锰", "2026-05-07", 6130],
                ["2026-05", "硅铁", "2026-05-07", 5810],
                ["2026-05", "低碳锰铁", "2026-05-07", 10150],
                ["2026-05", "铝块", "2026-05-07", 22000],
            ],
        )
    )

    assert report["status"] == "ok"
    assert report["parsed"]["tasks"][0]["manualAluminum"] == {
        "kgPerTon": pytest.approx(1.25),
        "pricePerTon": pytest.approx(22000),
        "materialName": "铝块",
    }
    assert report["parsed"]["tasks"][1]["manualAluminum"] == {
        "kgPerTon": 0,
        "pricePerTon": 0,
        "materialName": "",
    }


def test_single_target_values_follow_confirmed_exact_target_semantics_without_hidden_margins():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "C目标",
                "Si目标",
                "Mn目标",
                "P目标",
                "S目标",
                "V目标",
                "Nb目标",
                "Ti目标",
                "Als目标",
                "Alt目标",
                "Ca目标",
                "Cr目标",
                "Ni目标",
                "Cu目标",
                "Mo目标",
                "B目标",
                "Sb目标",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.16, 0.05, 0.20, 0.025, 0.020, None, None, None, 0.010, None, None, None, None, None, None, None, None],
                ["Q355C", 3, 12, "Q355C-1", 0.16, 0.10, 0.90, 0.018, 0.010, 0.030, 0.020, 0.025, None, 0.030, 0.004, 0.20, 0.10, 0.10, 0.05, 0.002, 0.03],
            ]
        )
    )

    assert report["status"] == "ok"
    q235b_target = report["parsed"]["tasks"][0]["config"]["target"]
    q235b_spec = report["parsed"]["tasks"][0]["config"]["target_spec"]
    q355c_target = report["parsed"]["tasks"][1]["config"]["target"]
    q355c_spec = report["parsed"]["tasks"][1]["config"]["target_spec"]

    assert q235b_spec["C"] == {"mode": "single", "value": 0.16}
    assert q235b_target["C"] == {"max": pytest.approx(0.155)}
    assert q235b_target["Si"] == {"max": 0.05}
    assert q235b_target["Mn"] == {"min": 0.20, "max": 0.20}
    assert q235b_target["P"] == {"max": 0.025}
    assert q235b_target["S"] == {"max": 0.020}
    assert q235b_spec["Als"] == {"mode": "single", "value": 0.010}
    assert "Als" not in q235b_target
    assert "Ca" not in q235b_target
    assert q355c_spec["Si"] == {"mode": "single", "value": 0.10}
    assert q355c_target["C"] == {"max": pytest.approx(0.155)}
    assert q355c_target["Si"] == {"min": 0.10, "max": 0.10}
    assert q355c_target["Mn"] == {"min": 0.90, "max": 0.90}
    assert q355c_target["P"] == {"max": 0.018}
    assert q355c_target["S"] == {"max": 0.010}
    assert q355c_target["V"] == {"min": 0.030, "max": 0.030}
    assert q355c_target["Nb"] == {"min": 0.020, "max": 0.020}
    assert q355c_spec["Ti"] == {"mode": "single", "value": 0.025}
    assert q355c_target["Ti"]["min"] == pytest.approx(0.030)
    assert q355c_target["Ti"]["max"] == pytest.approx(0.030)
    ti_bounds = effective_bounds(report["parsed"]["tasks"][1]["config"], "Ti")
    assert ti_bounds["min"] == pytest.approx(0.030)
    assert ti_bounds["max"] == pytest.approx(0.030)
    assert q355c_spec["Alt"] == {"mode": "single", "value": 0.030}
    assert "Alt" not in q355c_target
    assert "Ca" not in q355c_target
    assert q355c_target["Cr"] == {"min": 0.20, "max": 0.20}
    assert q355c_target["Ni"] == {"min": 0.10, "max": 0.10}
    assert q355c_target["Cu"] == {"min": 0.10, "max": 0.10}
    assert q355c_target["Mo"] == {"min": 0.05, "max": 0.05}
    assert q355c_target["B"] == {"min": 0.002, "max": 0.002}
    assert q355c_target["Sb"] == {"min": 0.03, "max": 0.03}


def test_rule_sheet_overrides_batch_process_rules():
    report = parse_template_workbook(
        workbook_bytes(
            rule_rows=[
                ["规则总开关", "enabled", "是", ""],
                ["控碳余量", "carbon_target_margin", 0.004, ""],
                ["禁硅阈值", "disable_silicon_alloys_si_max", 0.03, ""],
                ["铝块单独录入", "manual_aluminum", "否", ""],
                ["Ti 安全余量", "ti_safety_addition", 0.006, ""],
                ["Ni 禁投阈值", "trace_alloy_thresholds.Ni", 0.018, ""],
                ["P 禁投阈值", "phosphorus_alloy_max", 0.045, ""],
            ]
        )
    )

    assert report["status"] == "ok"
    rules = report["parsed"]["tasks"][0]["config"]["process_rules"]
    assert rules["enabled"] is True
    assert rules["carbon_target_margin"] == pytest.approx(0.004)
    assert rules["disable_silicon_alloys_si_max"] == pytest.approx(0.03)
    assert rules["single_target_si_upper_only_max"] == pytest.approx(0.05)
    assert rules["manual_aluminum"] is False
    assert rules["ti_safety_addition"] == pytest.approx(0.006)
    assert rules["trace_alloy_thresholds"]["Ni"] == pytest.approx(0.018)
    assert rules["phosphorus_alloy_max"] == pytest.approx(0.045)
    assert report["parsed"]["tasks"][0]["config"]["target"]["C"] == {"max": pytest.approx(0.156)}
    assert effective_bounds(report["parsed"]["tasks"][0]["config"], "C") == {"min": None, "max": pytest.approx(0.156)}


def test_rule_sheet_enabled_false_disables_batch_process_rules():
    report = parse_template_workbook(
        workbook_bytes(
            rule_rows=[
                ["规则总开关", "enabled", "否", ""],
                ["控碳余量", "carbon_target_margin", 0.004, ""],
                ["铝块单独录入", "manual_aluminum", "是", ""],
                ["Ti 安全余量", "ti_safety_addition", 0.006, ""],
            ]
        )
    )

    assert report["status"] == "ok"
    config = report["parsed"]["tasks"][0]["config"]
    assert config["process_rules"]["enabled"] is False
    assert config["target"]["C"] == {"max": pytest.approx(0.160)}
    assert effective_bounds(config, "C") == {"min": None, "max": pytest.approx(0.160)}


def test_non_element_target_headers_are_ignored():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "C目标",
                "Si目标",
                "Mn目标",
                "P目标",
                "S目标",
                "成本目标",
                "产量目标",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.16, 0.05, 0.20, 0.025, 0.020, 6100, 150],
                ["Q355C", 3, 12, "Q355C-1", 0.16, 0.10, 0.90, 0.018, 0.010, 6500, 160],
            ],
        )
    )

    assert report["status"] == "ok"
    target = report["parsed"]["tasks"][0]["config"]["target"]
    assert "成本" not in target
    assert "产量" not in target


def test_zero_target_value_is_treated_as_unconstrained():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "Mn目标",
                "Mo目标",
                "P目标",
                "S目标",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0, 0, 0.025, 0.020],
                ["Q355C", 3, 12, "Q355C-1", 0.90, 0, 0.018, 0.010],
            ],
        )
    )

    assert report["status"] == "ok"
    q235b_config = report["parsed"]["tasks"][0]["config"]
    q355c_config = report["parsed"]["tasks"][1]["config"]
    assert q235b_config["target_spec"]["Mn"]["mode"] == "none"
    assert q235b_config["target_spec"]["Mo"]["mode"] == "none"
    assert "Mn" not in q235b_config["target"]
    assert "Mo" not in q235b_config["target"]
    assert q355c_config["target_spec"]["Mo"]["mode"] == "none"
    assert "Mo" not in q355c_config["target"]


def test_legacy_ca_target_bounds_are_ignored():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "C上限",
                "Ca下限",
                "Ca上限",
                "Si上限",
                "Mn下限",
                "Mn上限",
                "P上限",
                "S上限",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.16, 0.001, 0.004, 0.05, 0.20, 0.22, 0.025, 0.020],
                ["Q355C", 3, 12, "Q355C-1", 0.16, 0.001, 0.005, 0.12, 0.90, 0.92, 0.018, 0.010],
            ],
        )
    )

    assert report["status"] == "ok"
    assert "Ca" not in report["parsed"]["tasks"][0]["config"]["target"]
    assert "Ca" not in report["parsed"]["tasks"][1]["config"]["target"]


def test_ca_recovery_is_not_carried_into_solver_config():
    report = parse_template_workbook(workbook_bytes())

    assert report["status"] == "ok"
    assert "Ca" not in report["parsed"]["tasks"][0]["config"]["recovery_rates"]
    assert "Ca" not in report["parsed"]["tasks"][1]["config"]["recovery_rates"]


def test_ca_does_not_appear_in_batch_chemistry_checks():
    report = parse_template_workbook(workbook_bytes())
    result = run_batch_optimization(report["parsed"], solver_name="internal")

    assert result["status"] == "ok"
    checks = result["results"][0]["result"]["modes"]["milp"]["chemistryChecks"]
    assert "Ca" not in {check["element"] for check in checks}


def test_al_recovery_is_fixed_at_fifteen_percent_even_when_sheet_values_differ():
    report = parse_template_workbook(workbook_bytes())

    assert report["status"] == "ok"
    rates = report["parsed"]["tasks"][0]["config"]["recovery_rates"]
    assert rates["Als"] == 0.15
    assert rates["Alt"] == 0.15


def test_invalid_al_recovery_cells_are_ignored_because_al_is_fixed():
    report = parse_template_workbook(
        workbook_bytes(
            endpoint_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.07, 0.12, 0, 0.90, 1.0, 1.0, 0.75, 0.98, 0.98, 0.98, 0.70, "不参与", "也不参与", 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["Q355C", 3, 12, "Q355C-1", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, None, None, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ]
        )
    )

    assert report["status"] == "ok"
    rates = report["parsed"]["tasks"][0]["config"]["recovery_rates"]
    assert rates["Als"] == 0.15
    assert rates["Alt"] == 0.15


def test_missing_cps_recovery_rates_are_reported():
    report = parse_template_workbook(
        workbook_bytes(
            endpoint_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "C终点",
                "Mn终点",
                "Cr终点",
                "Si回收率",
                "Mn回收率",
                "V回收率",
                "Nb回收率",
                "Ti回收率",
                "Als回收率",
                "Alt回收率",
                "Ca回收率",
                "Cr回收率",
                "Ni回收率",
                "Cu回收率",
                "Mo回收率",
                "B回收率",
                "Sb回收率",
            ],
            endpoint_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.07, 0.12, 0, 0.75, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["Q355C", 3, 12, "Q355C-1", 0.07, 0.11, 0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ],
        )
    )

    assert report["status"] == "error"
    missing_fields = {error["field"] for error in report["errors"] if error["code"] == "RECOVERY_NOT_FOUND"}
    assert {"C回收率", "P回收率", "S回收率"} <= missing_fields


def test_recovery_rates_are_matched_per_steelmaking_grade():
    report = parse_template_workbook(workbook_bytes())

    assert report["status"] == "ok"
    tasks = report["parsed"]["tasks"]
    assert tasks[0]["config"]["recovery_rates"]["Si"] == 0.75
    assert tasks[1]["config"]["recovery_rates"]["Si"] == 0.80
    assert abs(alloy_coeff(tasks[0]["config"]["alloys"][0], "Si", tasks[0]["config"]) - 17.69 * 0.75 / 1000) < 1e-12
    assert abs(alloy_coeff(tasks[1]["config"]["alloys"][0], "Si", tasks[1]["config"]) - 17.69 * 0.80 / 1000) < 1e-12


def test_26mnb5_zero_si_recovery_is_corrected_to_field_confirmed_value():
    report = parse_template_workbook(
        workbook_bytes(
            task_rows=[
                ["T001", "26MnB5", 10, 150, "2026-05", "26MnB5", "源表硅回收率 0 是录入错误"],
            ],
            target_rows=[
                ["26MnB5", 1.5, 12, "26MnB5", None, 0.26, 0.23, 0.25, 1.25, 1.27, 0.025, 0.020],
            ],
            endpoint_rows=[
                ["26MnB5", 1.5, 12, "26MnB5", 0.10, 0.0, 0, 0.90, 1.0, 1.0, 0.0, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ],
        )
    )

    assert report["status"] == "ok"
    assert report["parsed"]["tasks"][0]["config"]["recovery_rates"]["Si"] == 0.8


def test_single_target_and_legacy_bounds_for_same_element_conflict():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "Mn目标",
                "Mn下限",
                "Mn上限",
                "P目标",
                "S目标",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.20, 0.10, 0.30, 0.025, 0.020],
                ["Q355C", 3, 12, "Q355C-1", 0.90, None, None, 0.018, 0.010],
            ],
        )
    )

    assert report["status"] == "error"
    assert any(error["code"] == "TARGET_COLUMN_CONFLICT" and error["field"] == "Mn" for error in report["errors"])


def test_target_and_endpoint_steelmaking_grade_must_be_unique_across_sheet():
    report = parse_template_workbook(
        workbook_bytes(
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
                ["Q355C", 3, 12, "Q355C-1", None, 0.16, 0.10, 0.12, 0.90, 0.92, 0.018, 0.010],
                ["未引用重复", 1, 2, "UNUSED-DUP", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
                ["未引用重复2", 3, 4, "UNUSED-DUP", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
            ],
            endpoint_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.07, 0.12, 0, 0.90, 1.0, 1.0, 0.75, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["Q355C", 3, 12, "Q355C-1", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["未引用重复", 1, 2, "UNUSED-ENDPOINT-DUP", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["未引用重复2", 3, 4, "UNUSED-ENDPOINT-DUP", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ],
        )
    )

    assert report["status"] == "error"
    assert any(error["code"] == "DUPLICATE_STEELMAKING_GRADE" and error["sheet"] == "02_目标成分上下限" for error in report["errors"])
    assert any(error["code"] == "DUPLICATE_STEELMAKING_GRADE" and error["sheet"] == "03_转炉终点与回收率" for error in report["errors"])


def test_business_rows_match_only_by_unique_steelmaking_grade():
    report = parse_template_workbook(
        workbook_bytes(
            task_rows=[
                ["T001", "外部牌号名可不同", 999, 150, "2026-05", "Q235B-1", "只靠炼钢牌号定位"],
                ["T002", "Q355C", 8, 150, "2026-05", "Q355C-1", "低合金样例"],
            ]
        )
    )

    assert report["status"] == "ok"
    assert report["parsed"]["tasks"][0]["config"]["target"]["C"] == {"max": pytest.approx(0.155)}


def test_target_and_endpoint_legacy_match_fields_are_optional_when_steelmaking_grade_is_unique():
    report = parse_template_workbook(
        workbook_bytes(
            task_rows=[
                ["T001", "外部牌号名可不同", 999, 150, "2026-05", "Q235B-1", "只靠炼钢牌号定位"],
                ["T002", "Q355C", 8, 150, "2026-05", "Q355C-1", "低合金样例"],
            ],
            target_rows=[
                [None, None, None, "Q235B-1", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
                ["乱填也不参与匹配", 999, 1000, "Q355C-1", None, 0.16, 0.10, 0.12, 0.90, 0.92, 0.018, 0.010],
            ],
            endpoint_rows=[
                [None, None, None, "Q235B-1", 0.07, 0.12, 0, 0.90, 1.0, 1.0, 0.75, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["乱填也不参与匹配", 999, 1000, "Q355C-1", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ],
        )
    )

    assert report["status"] == "ok"
    assert report["parsed"]["tasks"][0]["config"]["target"]["Mn"] == {"min": 0.20, "max": 0.22}
    assert report["parsed"]["tasks"][0]["config"]["residual"]["C"] == 0.07


def test_target_and_endpoint_legacy_match_fields_do_not_validate_bad_values():
    report = parse_template_workbook(
        workbook_bytes(
            target_rows=[
                ["", "不是数字", "也不是数字", "Q235B-1", None, 0.16, None, 0.05, 0.20, 0.22, 0.025, 0.020],
                ["乱填也不参与匹配", 1000, 1, "Q355C-1", None, 0.16, 0.10, 0.12, 0.90, 0.92, 0.018, 0.010],
            ],
            endpoint_rows=[
                ["", "不是数字", "也不是数字", "Q235B-1", 0.07, 0.12, 0, 0.90, 1.0, 1.0, 0.75, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
                ["乱填也不参与匹配", 1000, 1, "Q355C-1", 0.07, 0.11, 0, 0.90, 1.0, 1.0, 0.80, 0.98, 0.98, 0.98, 0.70, 1.0, 1.0, 1.0, 0.96, 0.98, 0.98, 0.98, 0.70, 0.96],
            ],
        )
    )

    assert report["status"] == "ok"
    assert report["errors"] == []


def test_legacy_bound_target_columns_still_parse():
    report = parse_template_workbook(
        workbook_bytes(
            target_header=[
                "适用牌号",
                "最小厚度mm",
                "最大厚度mm",
                "炼钢牌号",
                "C下限",
                "C上限",
                "Si下限",
                "Si上限",
                "Mn下限",
                "Mn上限",
                "P上限",
                "S上限",
            ],
            target_rows=[
                ["Q235B", 1.5, 12, "Q235B-1", 0.06, 0.16, 0.05, 0.12, 0.20, 0.40, 0.025, 0.020],
                ["Q355C", 3, 12, "Q355C-1", 0.08, 0.16, 0.10, 0.20, 0.90, 1.20, 0.018, 0.010],
            ],
        )
    )

    assert report["status"] == "ok"
    assert report["parsed"]["tasks"][0]["config"]["target"]["Si"] == {"min": 0.05, "max": 0.12}


def test_batch_optimization_continues_when_one_task_fails():
    report = parse_template_workbook(workbook_bytes())
    parsed = report["parsed"]
    parsed["tasks"][1]["config"]["residual"]["P"] = 0.05

    result = run_batch_optimization(parsed, solver_name="internal")

    assert result["status"] == "partial_failed"
    assert result["summary"] == {"total": 2, "success": 1, "failed": 1}
    failed = next(item for item in result["results"] if item["taskId"] == "T002")
    assert failed["status"] == "failed"
    assert failed["errors"][0]["field"] == "P"


def test_template_validation_reports_duplicate_target_match():
    report = parse_template_workbook(workbook_bytes(duplicate_target=True))

    assert report["status"] == "error"
    assert any(error["code"] == "DUPLICATE_STEELMAKING_GRADE" and error["sheet"] == "02_目标成分上下限" for error in report["errors"])


def test_template_validation_rejects_merged_cells_and_formula_inputs():
    merged = parse_template_workbook(workbook_bytes(merge_tasks=True))
    formula = parse_template_workbook(workbook_bytes(formula_price=True))

    assert any(error["code"] == "MERGED_CELLS_NOT_ALLOWED" for error in merged["errors"])
    assert any(error["code"] == "FORMULA_NOT_ALLOWED" and error["sheet"] == "05_价格表" for error in formula["errors"])


def test_template_api_validate_batch_and_export():
    client = TestClient(app)

    validate = client.post(
        "/api/template/validate",
        files={"file": ("template.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert validate.status_code == 200
    payload = validate.json()
    assert payload["status"] == "ok"

    batch = client.post("/api/batch-optimize", json={"template": payload["parsed"], "solver": "internal"})
    assert batch.status_code == 200
    batch_payload = batch.json()
    assert batch_payload["status"] == "ok"
    assert batch_payload["summary"]["success"] == 2

    export = client.get(f"/api/batch-result/{batch_payload['batchId']}/export")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    exported = openpyxl.load_workbook(BytesIO(export.content), data_only=True)
    assert {"批量计算汇总", "最优路线明细", "成分校核", "错误与警告", "规则参数"} <= set(exported.sheetnames)


def test_template_download_api_returns_parseable_standard_template():
    client = TestClient(app)

    response = client.get("/api/template/download")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
    assert {
        "01_批量任务",
        "02_目标成分上下限",
        "03_转炉终点与回收率",
        "04_合金成分库",
        "05_价格表",
        "06_填写说明与校验规则",
        RULES_SHEET,
    } <= set(workbook.sheetnames)
    report = parse_template_workbook(response.content)
    assert report["status"] == "ok"
    assert report["preview"]["taskCount"] == 2


def test_download_template_uses_feed_mode_without_addition_sequence():
    workbook = openpyxl.load_workbook(BytesIO(generate_template_workbook()), data_only=True)
    headers = [cell.value for cell in workbook["04_合金成分库"][1]]
    rules_text = "\n".join(str(row[1].value or "") for row in workbook["06_填写说明与校验规则"].iter_rows(min_row=1, max_col=2))

    assert headers[:6] == ["合金名称", "价格物料名", "启用", "投料方式", "袋重kg", "最大投加kg每t"]
    assert "投料方式" in headers
    assert "投加顺序" not in headers
    assert "连续" in rules_text
    assert "整袋" in rules_text
    assert "路线序号" in rules_text
    assert "真实投料顺序优化" not in rules_text


def test_download_template_uses_requested_element_scope():
    workbook = openpyxl.load_workbook(BytesIO(generate_template_workbook()), data_only=True)
    target_headers = [cell.value for cell in workbook["02_目标成分上下限"][1]]
    endpoint_headers = [cell.value for cell in workbook["03_转炉终点与回收率"][1]]
    alloy_headers = [cell.value for cell in workbook["04_合金成分库"][1]]
    rules_text = "\n".join(str(row[1].value or "") for row in workbook["06_填写说明与校验规则"].iter_rows(min_row=1, max_col=2))

    assert target_headers[4:] == [f"{element}目标" for element in TEMPLATE_ELEMENTS]
    assert endpoint_headers[4 : 4 + len(TEMPLATE_ELEMENTS)] == [f"{element}终点" for element in TEMPLATE_ELEMENTS]
    assert endpoint_headers[4 + len(TEMPLATE_ELEMENTS) :] == [f"{element}回收率" for element in TEMPLATE_ELEMENTS]
    assert alloy_headers[6:-1] == TEMPLATE_ELEMENTS
    assert "N上限" not in target_headers
    assert "N回收率" not in endpoint_headers
    assert "N" not in alloy_headers
    assert "C/P/S 按上限控制" in rules_text
    assert "空值或 0 表示不约束" in rules_text
    assert "Si<=低硅阈值时只做低杂质控制、按上限处理" in rules_text
    assert "其余元素单值目标按精确值控制" in rules_text
    assert "Ti 单值=目标+余量" in rules_text
    assert "C目标-余量" in rules_text
    assert "铝块按现场单独录入" in rules_text
    assert "26MnB5 的 Si 回收率若录成 0" in rules_text
    assert "旧模板里的 N 上传时会被忽略" in rules_text
    rule_sheet = workbook[RULES_SHEET]
    rule_headers = [cell.value for cell in rule_sheet[1]]
    assert rule_headers == ["规则项", "参数键", "值", "说明"]
    assert rule_sheet["B2"].value == "enabled"
    assert rule_sheet["B5"].value == "single_target_si_upper_only_max"
    assert rule_sheet["B7"].value == "ti_safety_addition"


def test_checked_in_template_matches_generated_template_values():
    checked_in = openpyxl.load_workbook(Path("alloy-batch-template-v1.xlsx"), data_only=True)
    generated = openpyxl.load_workbook(BytesIO(generate_template_workbook()), data_only=True)

    assert checked_in.sheetnames == generated.sheetnames
    for sheet_name in generated.sheetnames:
        checked_sheet = checked_in[sheet_name]
        generated_sheet = generated[sheet_name]
        assert checked_sheet.max_row == generated_sheet.max_row
        assert checked_sheet.max_column == generated_sheet.max_column
        for row in range(1, generated_sheet.max_row + 1):
            for column in range(1, generated_sheet.max_column + 1):
                assert checked_sheet.cell(row, column).value == generated_sheet.cell(row, column).value


def test_feed_mode_continuous_allows_blank_or_zero_bag_size():
    rows = [
        ["硅锰", "硅锰", "是", "连续", None, 30, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, "空袋重连续"],
        ["硅铁", "硅铁", "是", "连续", 0, 20, 0.2, 72.23, None, None, 0.03, 0.02, None, "零袋重连续"],
        ["低碳锰铁", "低碳锰铁", "是", "整袋", 25, 25, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
    ]

    report = parse_template_workbook(workbook_bytes(alloy_rows=rows))

    assert report["status"] == "ok"
    alloys = report["parsed"]["tasks"][0]["config"]["alloys"]
    assert [alloy["feed_mode"] for alloy in alloys[:2]] == ["连续", "连续"]
    assert [alloy["bag_size_kg"] for alloy in alloys[:2]] == [0, 0]


def test_feed_mode_continuous_rejects_positive_bag_size():
    report = parse_template_workbook(
        workbook_bytes(
            alloy_rows=[
                ["硅锰", "硅锰", "是", "连续", 25, 30, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, ""],
                ["硅铁", "硅铁", "是", "连续", 0, 20, 0.2, 72.23, None, None, 0.03, 0.02, None, ""],
                ["低碳锰铁", "低碳锰铁", "是", "整袋", 25, 25, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
            ]
        )
    )

    assert report["status"] == "error"
    assert any(error["sheet"] == "04_合金成分库" and error["field"] == "袋重kg" for error in report["errors"])


def test_feed_mode_whole_bag_rejects_blank_or_zero_bag_size():
    for value in (None, 0):
        report = parse_template_workbook(
            workbook_bytes(
                alloy_rows=[
                    ["硅锰", "硅锰", "是", "连续", 0, 30, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, ""],
                    ["硅铁", "硅铁", "是", "连续", 0, 20, 0.2, 72.23, None, None, 0.03, 0.02, None, ""],
                    ["低碳锰铁", "低碳锰铁", "是", "整袋", value, 25, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
                ]
            )
        )

        assert report["status"] == "error"
        assert any(error["sheet"] == "04_合金成分库" and error["field"] == "袋重kg" for error in report["errors"])


def test_legacy_alloy_template_infers_feed_mode_and_warns():
    report = parse_template_workbook(
        workbook_bytes(
            alloy_header=["合金名称", "价格物料名", "启用", "袋重kg", "最大投加kg每t", "投加顺序", "C", "Si", "Mn", "Cr", "P", "S", "N", "备注"],
            alloy_rows=[
                ["硅锰", "硅锰", "是", 0, 30, 1, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, ""],
                ["硅铁", "硅铁", "是", None, 20, 2, 0.2, 72.23, None, None, 0.03, 0.02, None, ""],
                ["低碳锰铁", "低碳锰铁", "是", 25, 25, 3, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
            ],
        )
    )

    assert report["status"] == "ok"
    assert [alloy["feed_mode"] for alloy in report["parsed"]["tasks"][0]["config"]["alloys"]] == ["连续", "连续", "整袋"]
    assert any(warning["field"] == "投料方式" for warning in report["warnings"])
    assert any(warning["field"] == "投加顺序" for warning in report["warnings"])


def test_missing_feed_mode_without_legacy_sequence_is_structure_error():
    report = parse_template_workbook(
        workbook_bytes(
            alloy_header=["合金名称", "价格物料名", "启用", "袋重kg", "最大投加kg每t", "C", "Si", "Mn", "Cr", "P", "S", "N", "备注"],
            alloy_rows=[
                ["硅锰", "硅锰", "是", 0, 30, 1.72, 17.69, 65.66, None, 0.15, 0.02, 80, ""],
                ["硅铁", "硅铁", "是", 0, 20, 0.2, 72.23, None, None, 0.03, 0.02, None, ""],
                ["低碳锰铁", "低碳锰铁", "是", 25, 25, 0.64, None, 81.19, None, 0.20, 0.02, None, ""],
            ],
        )
    )

    assert report["status"] == "error"
    assert any(error["code"] == "MISSING_HEADER" and error["field"] == "投料方式" for error in report["errors"])


def test_batch_optimize_rejects_malformed_prevalidated_payload():
    client = TestClient(app)

    response = client.post("/api/batch-optimize", json={"template": {"tasks": [{"taskId": "T001"}]}, "solver": "internal"})

    assert response.status_code == 400
    assert "template 数据无效" in response.json()["detail"]


def test_batch_optimize_rejects_nested_malformed_prevalidated_payload():
    client = TestClient(app)

    response = client.post(
        "/api/batch-optimize",
        json={
            "solver": "internal",
            "template": {
                "tasks": [
                    {
                        "taskId": "T001",
                        "grade": "Q235B",
                        "thicknessMm": 10,
                        "config": {
                            "heat_weight_t": 150,
                            "target": [1],
                            "residual": {},
                            "recovery_rates": {},
                            "safety_margins": {},
                            "control_targets": {},
                            "process_rules": {},
                            "alloys": [1],
                            "milp_settings": {},
                        },
                    }
                ]
            },
        },
    )

    assert response.status_code == 400
    assert "config.target" in response.json()["detail"]


def test_export_route_details_filters_and_sorts_by_cost_without_addition_sequence():
    content = export_batch_result(
        {
            "results": [
                {
                    "status": "ok",
                    "input": {"taskId": "T001", "grade": "Q235B", "thicknessMm": 10},
                    "result": {
                        "modes": {
                            "milp": {
                                "costPerTon": 100,
                                "heatCost": 15000,
                                "alloys": [
                                    {"name": "低成本", "kgPerTon": 0, "heatKg": 0, "bags": 0, "costPerTon": 99, "sequence": 9, "feedMode": "连续"},
                                    {"name": "B合金", "kgPerTon": 1.5, "heatKg": 225, "bags": 9, "costPerTon": 20, "sequence": 3, "feedMode": "整袋"},
                                    {"name": "A合金", "kgPerTon": 2.0, "heatKg": 300, "bags": 0, "costPerTon": 20, "sequence": 1, "feedMode": "连续"},
                                    {"name": "高成本", "kgPerTon": 0.5, "heatKg": 75, "bags": 0, "costPerTon": 30, "sequence": 2, "feedMode": "连续"},
                                ],
                                "chemistryChecks": [],
                            }
                        },
                        "warnings": [],
                    },
                }
            ]
        }
    )
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    details = workbook["最优路线明细"]
    header = [cell.value for cell in details[1]]
    rows = [row for row in details.iter_rows(min_row=2, values_only=True)]

    assert header == ["任务编号", "路线序号", "合金", "kg/t", "炉次kg", "袋数", "成本贡献", "投料方式"]
    assert "投加顺序" not in header
    assert [(row[1], row[2], row[6], row[7]) for row in rows] == [
        (1, "高成本", 30, "连续"),
        (2, "A合金", 20, "连续"),
        (3, "B合金", 20, "整袋"),
    ]


def test_export_batch_result_records_manual_aluminum_without_adding_to_totals():
    content = export_batch_result(
        {
            "results": [
                {
                    "status": "ok",
                    "input": {
                        "taskId": "T001",
                        "grade": "Q235B",
                        "thicknessMm": 10,
                        "manualAluminum": {"kgPerTon": 1.5, "pricePerTon": 22000, "materialName": "铝块"},
                        "config": {"heat_weight_t": 150},
                    },
                    "result": {
                        "modes": {
                            "milp": {
                                "costPerTon": 100,
                                "heatCost": 15000,
                                "totalKgPerTon": 2,
                                "alloys": [
                                    {"name": "A合金", "kgPerTon": 2.0, "heatKg": 300, "bags": 0, "costPerTon": 100, "sequence": 1, "feedMode": "连续"},
                                ],
                                "chemistryChecks": [],
                            }
                        },
                        "warnings": [],
                    },
                }
            ]
        }
    )
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    summary = workbook["批量计算汇总"]
    details = workbook["最优路线明细"]
    summary_header = [cell.value for cell in summary[1]]
    summary_row = dict(zip(summary_header, [cell.value for cell in summary[2]]))
    detail_rows = [row for row in details.iter_rows(min_row=2, values_only=True)]

    assert summary_row["自动合金消耗kg/t"] == pytest.approx(2)
    assert summary_row["手工铝块kg/t"] == pytest.approx(1.5)
    assert summary_row["总合金消耗kg/t"] == pytest.approx(2)
    assert summary_row["自动吨钢成本"] == pytest.approx(100)
    assert summary_row["手工铝块成本元/t"] == pytest.approx(33)
    assert summary_row["总吨钢成本"] == pytest.approx(100)
    assert summary_row["炉次总成本"] == pytest.approx(15000)
    assert ("T001", 2, "铝块", 1.5, 225, None, 33, "手工录入") in detail_rows
