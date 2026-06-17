from pathlib import Path

import openpyxl

from tools.recalculate_lp_actual_aluminum import write_workbook


def test_recalculation_workbook_writes_composition_result_sheet(tmp_path):
    output = tmp_path / "comparison.xlsx"
    alloys = [
        {
            "name": "硅锰",
            "price_per_ton": 5000,
            "source_column": "AB",
            "composition": {"C": 1.72, "Si": 17.69, "Mn": 65.66},
        }
    ]
    rows = [
        {
            "Excel行号": 5,
            "序号": 1,
            "适用牌号": "样例",
            "分类序号": "A",
            "最小厚度mm": 1.5,
            "最大厚度mm": 2.0,
            "炼钢牌号": "SPHC",
            "钢种类型": "低碳",
            "液相线温度℃": 1510,
            "目标": {"C": 0.04, "Si": 0.02, "Mn": 0.20},
            "成分边界": {
                "C": {"min": None, "max": 0.035},
                "Si": {"min": None, "max": 0.02},
                "Mn": {"min": 0.20, "max": 0.20},
            },
            "终点成分": {"C": 0.025, "Si": 0, "Mn": 0.10},
            "LP状态": "ok",
            "Excel结果状态": "ok",
            "Excel是否满足批准规则": "否",
            "Excel原方案成分校核": "C=0.04(NG); Si=0.01(OK); Mn=0.20(OK)",
            "Excel不满足批准规则原因": "C高于上限：0.04% > 0.035%",
            "源表审计状态": "OK",
            "源表审计提示": "",
            "Excel合金消耗kg/t": 1.0,
            "Excel合金消耗原值": 1.0,
            "新算法合金消耗kg/t": 0.5,
            "新-Excel消耗kg/t": -0.5,
            "Excel合金成本元/t": 5.0,
            "Excel合金成本原值": 5.0,
            "新算法合金成本元/t": 2.5,
            "新-Excel成本元/t": -2.5,
            "Excel铝块kg/t": 0,
            "同源铝耗kg/t": 0,
            "铝耗来源": "1.合金成本!AH5",
            "铝耗口径": "同一工作簿AH列",
            "铝耗警告": "",
            "LP自动合金消耗kg/t(不含铝)": 0.5,
            "LP自动成本元/t(不含铝)": 2.5,
            "LP目标约束": "C<=0.035; Si<=0.02; Mn[0.2,0.2]",
            "新方案成分校核": "C=0.033(OK); Si=0.01(OK); Mn=0.20(OK)",
            "原因": "样例",
            "诊断": "",
            "old_x": [1.0],
            "old_raw_x": [1.0],
            "new_x": [0.5],
            "old_chemistry": {"C": 0.040, "Si": 0.010, "Mn": 0.200},
            "new_chemistry": {"C": 0.033, "Si": 0.010, "Mn": 0.200},
        }
    ]
    meta = {
        "summary": {
            "total": 1,
            "ok": 1,
            "infeasible": 0,
            "excel_result_errors": 0,
            "missing_aluminum": 0,
            "aluminum_warnings": 0,
            "recovery_overrides": 0,
            "source_audit_warnings": 0,
            "negative_alloy_rows": 0,
            "excel_rule_ok": 0,
            "excel_rule_ng": 1,
            "excel_rule_input_error": 0,
            "old_cost_total": 5.0,
            "new_cost_total": 2.5,
            "old_kg_total": 1.0,
            "new_kg_total": 0.5,
            "cost_delta_total": -2.5,
            "kg_delta_total": -0.5,
        },
        "alloys": alloys,
        "source_workbook": Path("source.xlsx"),
        "source_rows": [5],
        "process_rules": {
            "carbon_target_margin": 0.005,
            "disable_silicon_alloys_si_max": 0.04,
            "single_target_si_upper_only_max": 0.05,
            "ti_safety_addition": 0.005,
            "trace_alloy_thresholds": {"Ni": 0.02, "Cu": 0.02, "Mo": 0.02, "Sb": 0.02, "B": 0.0002},
            "phosphorus_alloy_max": 0.04,
            "sulfur_alloy_max": 0.03,
        },
    }

    write_workbook(rows, [], meta, output)

    workbook = openpyxl.load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["LP新算法铝耗对比", "成分结果"]
    sheet = workbook["成分结果"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[:12] == [
        "Excel行号",
        "炼钢牌号",
        "元素",
        "目标值",
        "约束下限",
        "约束上限",
        "终点成分",
        "Excel方案成分",
        "新算法成分",
        "新-Excel成分",
        "Excel校核",
        "新算法校核",
    ]
    rows_by_element = {sheet.cell(row, 3).value: row for row in range(2, sheet.max_row + 1)}
    assert rows_by_element.keys() >= {"C", "Si", "Mn"}
    c_row = rows_by_element["C"]
    assert sheet.cell(c_row, 10).value == -0.007
    assert sheet.cell(c_row, 11).value == "NG"
    assert sheet.cell(c_row, 12).value == "OK"
