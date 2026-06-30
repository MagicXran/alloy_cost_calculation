"""批量合金成本模板解析、校验、求解和导出。"""

from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from uuid import uuid4
from zipfile import BadZipFile

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from app.core import OptimizerError, solve_alloy_cost
from app.rules_engine import compile_rule_view, default_process_rules as load_default_process_rules, rule_template_allowed_keys, rule_template_rows
from app.solvers import get_solver


TEMPLATE_VERSION = "1"
IGNORED_ELEMENTS = {"N"}
TEMPLATE_ELEMENTS = ["C", "Si", "Mn", "P", "S", "V", "Nb", "Ti", "Als", "Alt", "Ca", "Cr", "Ni", "Cu", "Mo", "B", "Sb"]
ERROR_FIELD_ELEMENTS = ["P", "S"] + [element for element in TEMPLATE_ELEMENTS if element not in {"P", "S"}]
TARGET_IGNORED_ELEMENTS = {"Ca"}
FIXED_RECOVERY_RATES = {"Als": 0.15, "Alt": 0.15}
FIELD_CONFIRMED_RECOVERY_OVERRIDES = {("26MNB5", "Si"): 0.8}
MANUAL_ALUMINUM_FIELD = "手工铝块kg/t"
ALUMINUM_ALIASES = ("铝块", "铝粒", "铝锭", "铝线")
BUSINESS_SHEETS = [
    "01_批量任务",
    "02_目标成分上下限",
    "03_转炉终点与回收率",
    "04_合金成分库",
    "05_价格表",
]
REQUIRED_SHEETS = BUSINESS_SHEETS + ["06_填写说明与校验规则"]
REQUIRED_HEADERS = {
    "01_批量任务": ["任务编号", "适用牌号", "厚度mm", "炉重t", "价格方案", "炼钢牌号"],
    "02_目标成分上下限": ["炼钢牌号"],
    "03_转炉终点与回收率": ["炼钢牌号"],
    "04_合金成分库": ["合金名称", "价格物料名", "启用", "袋重kg", "最大投加kg每t"],
    "05_价格表": ["价格方案", "物料名称", "价格日期", "价格元每吨"],
}
RULES_SHEET = "07_工艺规则参数"
RULES_HEADERS = ["规则项", "参数键", "值", "说明"]
ALLOY_METADATA_HEADERS = set(REQUIRED_HEADERS["04_合金成分库"]) | {"投料方式", "投加顺序", "备注"}
FEED_MODES = {"连续", "整袋"}
BATCH_RESULTS: dict[str, dict] = {}


def default_process_rules() -> dict:
    """返回模板与批量链路使用的默认工艺规则。"""

    return load_default_process_rules()


def rule_template_value(process_rules: dict, key: str):
    """从嵌套规则 dict 读取模板展示值。"""

    current = process_rules
    for part in key.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def generate_template_workbook(process_rules: dict | None = None) -> bytes:
    """生成系统推荐的批量计算上传模板。"""

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    template_rules = deepcopy(process_rules) if isinstance(process_rules, dict) else default_process_rules()

    q235b_target = {"C": 0.16, "Si": 0.05, "Mn": 0.20, "P": 0.025, "S": 0.020}
    q355c_target = {"C": 0.16, "Si": 0.10, "Mn": 0.90, "P": 0.018, "S": 0.010}
    q235b_residual = {"C": 0.07, "Mn": 0.12, "Cr": 0}
    q355c_residual = {"C": 0.07, "Mn": 0.11, "Cr": 0}
    q235b_recovery = {**FIXED_RECOVERY_RATES, "C": 0.90, "Si": 0.75, "Mn": 0.98, "P": 1.00, "S": 1.00, "Cr": 0.96}
    q355c_recovery = {**FIXED_RECOVERY_RATES, "C": 0.90, "Si": 0.80, "Mn": 0.98, "P": 1.00, "S": 1.00, "Cr": 0.96}

    create_sheet(
        workbook,
        "01_批量任务",
        [
            ["任务编号", "适用牌号", "厚度mm", "炉重t", "价格方案", "炼钢牌号", MANUAL_ALUMINUM_FIELD, "备注"],
            ["T001", "Q235B", 10, 150, "2026-05", "Q235B-1", None, "正常样例，可复制新增任务"],
            ["T002", "Q355C", 8, 150, "2026-05", "Q355C-1", None, "批量任务互不影响"],
        ],
        {"A": 16, "B": 14, "C": 12, "D": 12, "E": 14, "F": 16, "G": 16, "H": 28},
    )
    create_sheet(
        workbook,
        "02_目标成分上下限",
        [
            ["适用牌号", "最小厚度mm", "最大厚度mm", "炼钢牌号", *element_target_headers()],
            ["Q235B", 1.5, 12, "Q235B-1", *element_values(q235b_target)],
            ["Q355C", 3, 12, "Q355C-1", *element_values(q355c_target)],
        ],
        {"A": 14, "B": 14, "C": 14, "D": 16},
    )
    create_sheet(
        workbook,
        "03_转炉终点与回收率",
        [
            ["适用牌号", "最小厚度mm", "最大厚度mm", "炼钢牌号", *element_endpoint_headers(), *element_recovery_headers()],
            ["Q235B", 1.5, 12, "Q235B-1", *element_values(q235b_residual), *element_values(q235b_recovery)],
            ["Q355C", 3, 12, "Q355C-1", *element_values(q355c_residual), *element_values(q355c_recovery)],
        ],
        {"A": 14, "B": 14, "C": 14, "D": 16},
    )
    create_sheet(
        workbook,
        "04_合金成分库",
        [
            ["合金名称", "价格物料名", "启用", "投料方式", "袋重kg", "最大投加kg每t", *TEMPLATE_ELEMENTS, "备注"],
            ["硅锰", "硅锰", "是", "连续", 0, 30, *element_values({"C": 1.72, "Si": 17.69, "Mn": 65.66, "P": 0.15, "S": 0.02}), ""],
            ["高碳锰铁", "高碳锰铁", "是", "整袋", 25, 25, *element_values({"C": 6.69, "Mn": 74.60, "P": 0.20, "S": 0.02}), ""],
            ["中碳锰铁", "中碳锰铁", "是", "整袋", 25, 25, *element_values({"C": 1.50, "Mn": 78.54, "P": 0.20, "S": 0.02}), ""],
            ["低碳锰铁", "低碳锰铁", "是", "整袋", 25, 25, *element_values({"C": 0.64, "Mn": 81.19, "P": 0.20, "S": 0.02}), ""],
            ["硅铁", "硅铁", "是", "连续", 0, 20, *element_values({"C": 0.20, "Si": 72.23, "P": 0.03, "S": 0.02}), ""],
            ["高碳铬铁", "高碳铬铁", "是", "连续", 0, 20, *element_values({"C": 10.00, "P": 0.03, "S": 0.04, "Cr": 52.00}), ""],
            ["低碳铬铁", "低碳铬铁", "是", "连续", 0, 20, *element_values({"C": 0.12, "P": 0.03, "S": 0.03, "Cr": 59.44}), ""],
            ["铝块", "铝块", "是", "连续", 0, 10, *element_values({"P": 0.00, "S": 0.00, "Als": 99.00, "Alt": 99.00}), "手工铝块按 01_批量任务 的手工铝块kg/t 单独记录，不参与 LP 自动优化，也不计入最终合金成本和消耗。"],
        ],
        {"A": 14, "B": 14, "C": 10, "D": 12, "E": 10, "F": 16, "W": 28},
    )
    create_sheet(
        workbook,
        "05_价格表",
        [
            ["价格方案", "物料名称", "价格日期", "价格元每吨"],
            ["2026-05", "硅锰", "2026-05-07", 5088],
            ["2026-05", "高碳锰铁", "2026-05-07", 5593],
            ["2026-05", "中碳锰铁", "2026-05-07", 7420],
            ["2026-05", "低碳锰铁", "2026-05-07", 7876],
            ["2026-05", "硅铁", "2026-05-07", 4973],
            ["2026-05", "高碳铬铁", "2026-05-07", 7699],
            ["2026-05", "低碳铬铁", "2026-05-07", 14956],
            ["2026-05", "铝块", "2026-05-07", 22337],
        ],
        {"A": 14, "B": 14, "C": 14, "D": 14},
    )
    create_sheet(
        workbook,
        "06_填写说明与校验规则",
        [
            ["模板版本", TEMPLATE_VERSION],
            ["填写流程", "下载模板 -> 填业务数据 -> 上传预检 -> 预检通过后批量计算 -> 导出结果"],
            ["单位规则", "成分按百分数数值填写，例如 0.23 表示 0.23%；合金品位 65.66 表示 65.66%。"],
            ["外部单值规则", "目标成分表使用 元素目标 单值列：空值或 0 表示不约束；C/P/S 按上限控制；Si<=低硅阈值时只做低杂质控制、按上限处理；Ni/Cu/Mo/Sb<=对应阈值、B<=对应阈值时可留空或直接视为不投；其余元素单值目标按精确值控制。旧的 元素下限/元素上限 上传列仍兼容。"],
            ["现场工艺规则", "以 07_工艺规则参数 sheet 为准：当前批准规则包括 C目标-余量、Si<=阈值禁硅、金属锰兜底、铝块按现场单独录入、Ti 单值=目标+余量、Ti 区间下限+余量、Ni/Cu/Mo/Sb/B 低目标禁投、P/S 低目标禁投磷硫铁。"],
            ["手工铝块", "01_批量任务 可填写 手工铝块kg/t；该值不参与 LP 自动优化，也不计入最终合金成本和合金消耗。空值或 0 表示本炉没有手工铝块记录。"],
            ["合金用量公式", "kg/t = (目标成分 - 有效终点成分) / 合金品位 / 回收率 * 1000；当前批量链路只使用 C/Mn 终点扣减，V/Nb/Ti/Cr 等不做终点残余扣减；26MnB5 的 Si 回收率若录成 0 会按现场确认值 0.8 修正。"],
            ["标准元素", "标准模板仅保留 C, Si, Mn, P, S, V, Nb, Ti, Als, Alt, Ca, Cr, Ni, Cu, Mo, B, Sb；旧模板里的 N 上传时会被忽略。"],
            ["P/S 规则", "P/S 通常只填写上限；转炉终点已超过上限时任务直接失败。"],
            ["投料方式", "投料方式只能填写 连续 或 整袋；连续物料袋重kg留空或填 0，整袋物料袋重kg必须大于 0。"],
            ["路线序号", "最优路线明细中的路线序号只表示导出排序，从 1 开始，按成本贡献降序排列；它不是现场真实投料顺序。"],
            ["规则参数", "07_工艺规则参数 是可编辑业务 sheet；修改阈值后，预检通过的批量任务会按该 sheet 中的规则求解。"],
            ["禁止内容", "业务输入区不允许合并单元格、公式、空表头、重复表头、隐藏必填列。"],
            ["错误定位", "系统会返回 sheet、行号、字段、错误码、原因和修正建议。"],
        ],
        {"A": 18, "B": 96},
    )
    create_sheet(
        workbook,
        RULES_SHEET,
        [
            RULES_HEADERS,
            *[
                [label, key, rule_template_value(template_rules, key), note]
                for label, key, note in rule_template_rows()
            ],
        ],
        {"A": 18, "B": 28, "C": 14, "D": 70},
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_sheet(workbook, title: str, rows: list[list], widths: dict[str, int]) -> None:
    """创建一个无合并单元格的模板 sheet。"""

    sheet = workbook.create_sheet(title)
    for row in rows:
        sheet.append(row)
    style_template_sheet(sheet, widths)


def style_template_sheet(sheet, widths: dict[str, int]) -> None:
    """给模板 sheet 加基础可读样式，不引入公式和合并单元格。"""

    header_fill = PatternFill("solid", fgColor="0A0B0D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        width = widths.get(letter)
        if width is None:
            width = min(18, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[letter].width = width


def element_target_headers() -> list[str]:
    return [f"{element}目标" for element in TEMPLATE_ELEMENTS]


def element_endpoint_headers() -> list[str]:
    return [f"{element}终点" for element in TEMPLATE_ELEMENTS]


def element_recovery_headers() -> list[str]:
    return [f"{element}回收率" for element in TEMPLATE_ELEMENTS]


def element_values(values_by_element: dict[str, float]) -> list[float | None]:
    return [values_by_element.get(element) for element in TEMPLATE_ELEMENTS]


def parse_template_workbook(content: bytes) -> dict:
    """解析用户上传模板，并返回预检报告和标准化数据。"""

    errors: list[dict] = []
    warnings: list[dict] = []
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        return report_with_errors(
            [make_issue(None, None, None, "INVALID_XLSX", f"无法读取 xlsx 文件：{exc}", "请上传未损坏的 .xlsx 模板。")]
        )

    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    for sheet in missing:
        errors.append(make_issue(sheet, None, None, "MISSING_SHEET", f"缺少 sheet：{sheet}", "请使用系统模板，不要改 sheet 名。"))
    if errors:
        return report_with_errors(errors)

    rows_by_sheet: dict[str, list[dict]] = {}
    for sheet_name in BUSINESS_SHEETS:
        rows_by_sheet[sheet_name] = read_business_sheet(workbook[sheet_name], errors)
    if RULES_SHEET in workbook.sheetnames:
        rows_by_sheet[RULES_SHEET] = read_business_sheet(workbook[RULES_SHEET], errors, required_headers=RULES_HEADERS)

    if errors:
        return report_with_errors(errors)

    parsed = build_parsed_template(rows_by_sheet, errors, warnings)
    status = "ok" if not errors else "error"
    return {
        "status": status,
        "templateVersion": TEMPLATE_VERSION,
        "errors": errors,
        "warnings": warnings,
        "preview": {
            "taskCount": len(parsed.get("tasks", [])) if parsed else 0,
            "alloyCount": len(parsed.get("alloys", [])) if parsed else 0,
            "priceCount": len(parsed.get("prices", [])) if parsed else 0,
        },
        "parsed": parsed if status == "ok" else None,
    }


def report_with_errors(errors: list[dict]) -> dict:
    """返回统一错误报告。"""

    return {
        "status": "error",
        "templateVersion": TEMPLATE_VERSION,
        "errors": errors,
        "warnings": [],
        "preview": {"taskCount": 0, "alloyCount": 0, "priceCount": 0},
        "parsed": None,
    }


def read_business_sheet(sheet, errors: list[dict], required_headers: list[str] | None = None) -> list[dict]:
    """读取一个业务 sheet，并做结构级校验。"""

    sheet_name = sheet.title
    if sheet.merged_cells.ranges:
        errors.append(
            make_issue(
                sheet_name,
                None,
                None,
                "MERGED_CELLS_NOT_ALLOWED",
                "业务输入区不允许合并单元格。",
                "请取消合并单元格，每个字段只占一个单元格。",
            )
        )

    header_row = [normalize_header(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    seen: dict[str, int] = {}
    for index, header in enumerate(header_row, 1):
        if not header:
            continue
        if header in seen:
            errors.append(make_issue(sheet_name, 1, header, "DUPLICATE_HEADER", f"表头重复：{header}", "请删除或重命名重复表头。"))
        seen[header] = index

    for header in required_headers or REQUIRED_HEADERS[sheet_name]:
        column = seen.get(header)
        if column is None:
            errors.append(make_issue(sheet_name, 1, header, "MISSING_HEADER", f"缺少必填字段：{header}", "请使用系统模板表头。"))
            continue
        letter = sheet.cell(1, column).column_letter
        if sheet.column_dimensions[letter].hidden:
            errors.append(make_issue(sheet_name, 1, header, "HIDDEN_REQUIRED_COLUMN", f"必填列被隐藏：{header}", "请取消隐藏该列。"))

    if sheet_name == "04_合金成分库":
        feed_mode_column = seen.get("投料方式")
        legacy_sequence_column = seen.get("投加顺序")
        if feed_mode_column is None and legacy_sequence_column is None:
            errors.append(make_issue(sheet_name, 1, "投料方式", "MISSING_HEADER", "缺少必填字段：投料方式", "请使用新模板填写 连续/整袋；旧模板兼容仅限保留投加顺序字段的一版。"))
        elif feed_mode_column is not None:
            letter = sheet.cell(1, feed_mode_column).column_letter
            if sheet.column_dimensions[letter].hidden:
                errors.append(make_issue(sheet_name, 1, "投料方式", "HIDDEN_REQUIRED_COLUMN", "必填列被隐藏：投料方式", "请取消隐藏该列。"))

    rows: list[dict] = []
    for row_index in range(2, sheet.max_row + 1):
        if is_empty_row(sheet, row_index):
            continue
        item = {"_row": row_index}
        for column_index, header in enumerate(header_row, 1):
            if not header:
                continue
            cell = sheet.cell(row_index, column_index)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                errors.append(make_issue(sheet_name, row_index, header, "FORMULA_NOT_ALLOWED", "业务输入区不允许公式。", "请粘贴公式计算后的值。"))
            if isinstance(cell.value, str) and cell.value.strip().startswith("#"):
                errors.append(make_issue(sheet_name, row_index, header, "SPREADSHEET_ERROR", f"单元格包含错误值：{cell.value}", "请修正后重新上传。"))
            item[header] = cell.value
        rows.append(item)
    return rows


def build_parsed_template(rows_by_sheet: dict[str, list[dict]], errors: list[dict], warnings: list[dict]) -> dict:
    """把各 sheet 原始行转换成批量求解输入。"""

    process_rules = parse_process_rules_rows(rows_by_sheet.get(RULES_SHEET) or [], errors)
    target_rows = parse_target_rows(rows_by_sheet["02_目标成分上下限"], errors, process_rules)
    endpoint_rows = parse_endpoint_rows(rows_by_sheet["03_转炉终点与回收率"], errors)
    validate_unique_steelmaking_grade(target_rows, "02_目标成分上下限", errors)
    validate_unique_steelmaking_grade(endpoint_rows, "03_转炉终点与回收率", errors)
    alloy_rows = parse_alloy_rows(rows_by_sheet["04_合金成分库"], errors, warnings)
    prices = parse_price_rows(rows_by_sheet["05_价格表"], errors)
    task_rows = parse_task_rows(rows_by_sheet["01_批量任务"], errors)
    if errors:
        return {}

    price_map = {(item["priceScheme"], item["materialName"]): item for item in prices}
    parsed_tasks = []
    for task in task_rows:
        task_errors: list[dict] = []
        target = match_business_row(target_rows, task, "02_目标成分上下限", task_errors)
        endpoint = match_business_row(endpoint_rows, task, "03_转炉终点与回收率", task_errors)
        if target is None or endpoint is None:
            errors.extend(task_errors)
            continue

        config_alloys = []
        for alloy in alloy_rows:
            if not alloy["enabled"]:
                continue
            price = price_map.get((task["priceScheme"], alloy["priceMaterialName"]))
            if price is None:
                errors.append(
                    make_issue(
                        "05_价格表",
                        None,
                        "价格元每吨",
                        "PRICE_NOT_FOUND",
                        f"价格方案 {task['priceScheme']} 下找不到物料 {alloy['priceMaterialName']} 的价格。",
                        "请在价格表补充该物料价格，或修正合金成分库里的价格物料名。",
                    )
                )
                continue
            config_alloys.append(
                {
                    "name": alloy["name"],
                    "price_per_ton": price["pricePerTon"],
                    "composition": alloy["composition"],
                    "enabled": True,
                    "addition_sequence": 99,
                    "feed_mode": alloy["feedMode"],
                    "bag_size_kg": alloy["bagSizeKg"],
                    "max_add_kg_per_t": alloy["maxAddKgPerT"],
                    "recovery_overrides": {},
                    "notes": alloy.get("notes", ""),
                }
            )

        if not config_alloys:
            errors.append(make_issue("04_合金成分库", None, "启用", "NO_ENABLED_ALLOYS", "没有可参与计算的启用合金。", "请至少启用一种合金并配置价格。"))
            continue

        manual_aluminum = manual_aluminum_for_task(task, alloy_rows, price_map, errors)
        target_spec = deepcopy(target["targetSpec"])
        residual = build_residual(target_spec, endpoint["residual"], config_alloys)
        recovery_rates = build_recovery_rates(target_spec, endpoint["recoveryRates"], endpoint.get("row"), config_alloys, errors)
        preview_config = {
            "target_spec": deepcopy(target_spec),
            "process_rules": deepcopy(process_rules),
            "control_targets": {"enabled": False, "margin": 0, "elements": {}},
            "safety_margins": {element: {"low": 0, "high": 0} for element in target_spec},
            "alloys": config_alloys,
        }
        preview_view = compile_rule_view(preview_config)
        compiled_target = preview_view.as_legacy_target()
        parsed_tasks.append(
            {
                "taskId": task["taskId"],
                "grade": task["grade"],
                "thicknessMm": task["thicknessMm"],
                "priceScheme": task["priceScheme"],
                "row": task["row"],
                "manualAluminum": manual_aluminum,
                "config": {
                    "heat_weight_t": task["heatWeightT"],
                    "steel_weight_kg": 1000,
                    "ignored_elements": sorted(IGNORED_ELEMENTS),
                    "target_spec": deepcopy(preview_view.target_spec),
                    "target": compiled_target,
                    "residual": residual,
                    "recovery_rates": recovery_rates,
                    "safety_margins": {element: {"low": 0, "high": 0} for element in target_spec},
                    "control_targets": {"enabled": False, "margin": 0, "elements": {}},
                    "process_rules": deepcopy(process_rules),
                    "alloys": config_alloys,
                    "milp_settings": {"default_bag_size_kg": 25, "enable_bag_rounding": True},
                    "temperature_drop": {"enabled": False},
                },
            }
        )
    if errors:
        return {}
    return {"templateVersion": TEMPLATE_VERSION, "tasks": parsed_tasks, "alloys": alloy_rows, "prices": prices, "processRules": process_rules}


def manual_aluminum_for_task(task: dict, alloy_rows: list[dict], price_map: dict[tuple[str, str], dict], errors: list[dict]) -> dict:
    """返回逐炉手工铝块记录信息；空值兼容旧模板。"""

    kg_per_ton = float(task.get("manualAluminumKgPerT") or 0)
    if kg_per_ton <= 0:
        return {"kgPerTon": 0, "pricePerTon": 0, "materialName": ""}

    alloy = next((item for item in alloy_rows if is_aluminum_name(item.get("name", ""))), None)
    material_name = (alloy or {}).get("priceMaterialName") or "铝块"
    price = price_map.get((task["priceScheme"], material_name))
    if price is None:
        errors.append(
            make_issue(
                "05_价格表",
                None,
                "价格元每吨",
                "MANUAL_ALUMINUM_PRICE_NOT_FOUND",
                f"任务 {task['taskId']} 填写了手工铝块kg/t，但价格方案 {task['priceScheme']} 下找不到铝块物料 {material_name} 的价格。",
                "请在价格表补充铝块价格，或把手工铝块kg/t 留空/填 0。",
            )
        )
        return {"kgPerTon": kg_per_ton, "pricePerTon": 0, "materialName": material_name}
    return {"kgPerTon": kg_per_ton, "pricePerTon": price["pricePerTon"], "materialName": material_name}


def normalized_manual_aluminum(value: dict | None) -> dict:
    """规范化预检产物里的手工铝块对象，兼容旧 payload。"""

    if not isinstance(value, dict):
        return {"kgPerTon": 0.0, "pricePerTon": 0.0, "materialName": ""}
    return {
        "kgPerTon": float(value.get("kgPerTon") or 0),
        "pricePerTon": float(value.get("pricePerTon") or 0),
        "materialName": optional_text(value.get("materialName")) or "",
    }


def manual_aluminum_cost_per_ton(manual_aluminum: dict) -> float:
    return float(manual_aluminum.get("kgPerTon") or 0) * float(manual_aluminum.get("pricePerTon") or 0) / 1000


def is_aluminum_name(name: str) -> bool:
    normalized = str(name or "").replace(" ", "")
    return any(alias in normalized for alias in ALUMINUM_ALIASES)


def parse_process_rules_rows(rows: list[dict], errors: list[dict]) -> dict:
    """解析模板里的可编辑工艺规则；缺省时退回系统默认值。"""

    rules = default_process_rules()
    if not rows:
        return rules

    allowed_keys = rule_template_allowed_keys()
    seen: set[str] = set()
    for row in rows:
        key = optional_text(row.get("参数键"))
        if not key:
            continue
        if key in seen:
            errors.append(make_issue(RULES_SHEET, row["_row"], "参数键", "DUPLICATE_RULE_KEY", f"规则参数重复：{key}", "每个参数键只能出现一次。"))
            continue
        seen.add(key)
        if key not in allowed_keys:
            errors.append(make_issue(RULES_SHEET, row["_row"], "参数键", "UNKNOWN_RULE_KEY", f"未知规则参数：{key}", "请使用系统模板给出的参数键，不要自行改名。"))
            continue

        raw_value = row.get("值")
        if raw_value in (None, ""):
            continue

        if key in {"enabled", "manual_aluminum"}:
            parsed = parse_bool_flag(raw_value, RULES_SHEET, row["_row"], "值", errors)
            if parsed is None:
                continue
            rules[key] = parsed
            continue

        maximum = 10 if key in {"disable_silicon_alloys_si_max", "single_target_si_upper_only_max", "carbon_target_margin", "ti_safety_addition", "phosphorus_alloy_max", "sulfur_alloy_max"} else 1.2
        parsed_number = optional_number(raw_value, RULES_SHEET, row["_row"], "值", errors, minimum=0, maximum=maximum, allow_zero=True)
        if parsed_number is None:
            continue

        if key.startswith("trace_alloy_thresholds."):
            element = key.split(".", 1)[1]
            thresholds = dict(rules.get("trace_alloy_thresholds") or {})
            thresholds[element] = parsed_number
            rules["trace_alloy_thresholds"] = thresholds
        else:
            rules[key] = parsed_number
    return rules


def run_batch_optimization(parsed_template: dict, solver_name: str = "highs") -> dict:
    """逐行调用优化器，一行失败不阻断其他行。"""

    tasks = validate_prechecked_template(parsed_template)
    solver = get_solver(solver_name)
    results = []
    for task in tasks:
        try:
            result = solve_alloy_cost(task["config"], solver)
            if result.get("status") == "ok":
                results.append({"taskId": task["taskId"], "status": "ok", "input": task, "result": result, "errors": []})
            else:
                errors = [
                    make_issue("01_批量任务", task.get("row"), None, "OPTIMIZER_INFEASIBLE", "求解无可行解。", "请检查目标范围、残余成分、启用合金和袋重约束。")
                ]
                for diagnostic in result.get("diagnostics") or []:
                    errors.append(make_issue("01_批量任务", task.get("row"), None, "OPTIMIZER_DIAGNOSTIC", diagnostic, "按诊断调整模板数据。"))
                results.append({"taskId": task["taskId"], "status": "failed", "input": task, "result": result, "errors": errors})
        except OptimizerError as exc:
            results.append(
                {
                    "taskId": task["taskId"],
                    "status": "failed",
                    "input": task,
                    "result": None,
                    "errors": [optimizer_error_to_issue(task, exc)],
                }
            )
        except ValueError as exc:
            results.append({"taskId": task["taskId"], "status": "failed", "input": task, "result": None, "errors": [make_issue("01_批量任务", task.get("row"), None, "SOLVER_ERROR", str(exc), "请检查 solver 参数。")]})

    success = sum(1 for item in results if item["status"] == "ok")
    failed = len(results) - success
    batch_id = uuid4().hex
    payload = {
        "batchId": batch_id,
        "status": "ok" if failed == 0 else ("failed" if success == 0 else "partial_failed"),
        "summary": {"total": len(results), "success": success, "failed": failed},
        "results": results,
    }
    BATCH_RESULTS[batch_id] = payload
    return payload


def validate_prechecked_template(parsed_template: dict) -> list[dict]:
    """校验 /api/batch-optimize 收到的对象确实像预检产物，避免半截 JSON 触发 KeyError。"""

    if not isinstance(parsed_template, dict) or not isinstance(parsed_template.get("tasks"), list):
        raise ValueError("template 数据无效，请先调用 /api/template/validate 并提交返回的 parsed。")
    tasks = parsed_template["tasks"]
    for index, task in enumerate(tasks, 1):
        if not isinstance(task, dict):
            raise ValueError(f"template 数据无效：第 {index} 个任务不是对象，请先调用 /api/template/validate 并提交返回的 parsed。")
        missing = [field for field in ("taskId", "grade", "thicknessMm", "config") if field not in task]
        if missing or not isinstance(task.get("config"), dict):
            missing_text = "、".join(missing or ["config"])
            raise ValueError(f"template 数据无效：第 {index} 个任务缺少 {missing_text}，请先调用 /api/template/validate 并提交返回的 parsed。")
        if "manualAluminum" in task and not isinstance(task.get("manualAluminum"), dict):
            raise ValueError(f"template 数据无效：第 {index} 个任务的 manualAluminum 类型错误，请先调用 /api/template/validate 并提交返回的 parsed。")
        validate_prechecked_config(task["config"], index)
    return tasks


def validate_prechecked_config(config: dict, task_index: int) -> None:
    """只做结构型校验；数值和业务可行性仍交给优化器逐行处理。"""

    expected_types = {
        "target": dict,
        "target_spec": dict,
        "residual": dict,
        "recovery_rates": dict,
        "safety_margins": dict,
        "control_targets": dict,
        "process_rules": dict,
        "alloys": list,
        "milp_settings": dict,
    }
    for field, expected_type in expected_types.items():
        if not isinstance(config.get(field), expected_type):
            raise ValueError(f"template 数据无效：第 {task_index} 个任务的 config.{field} 类型错误，请先调用 /api/template/validate 并提交返回的 parsed。")
    for alloy_index, alloy in enumerate(config["alloys"], 1):
        if not isinstance(alloy, dict):
            raise ValueError(f"template 数据无效：第 {task_index} 个任务的 config.alloys[{alloy_index}] 类型错误，请先调用 /api/template/validate 并提交返回的 parsed。")


def export_batch_result(batch_result: dict) -> bytes:
    """把批量计算结果导出为 Excel 字节。"""

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "批量计算汇总"
    details = workbook.create_sheet("最优路线明细")
    chemistry = workbook.create_sheet("成分校核")
    issues = workbook.create_sheet("错误与警告")
    rule_sheet = workbook.create_sheet("规则参数")

    summary.append(
        [
            "任务编号",
            "牌号",
            "厚度mm",
            "状态",
            "自动吨钢成本",
            "手工铝块成本元/t",
            "总吨钢成本",
            "自动合金消耗kg/t",
            "手工铝块kg/t",
            "总合金消耗kg/t",
            "炉次总成本",
            "失败原因",
        ]
    )
    details.append(["任务编号", "路线序号", "合金", "kg/t", "炉次kg", "袋数", "成本贡献", "投料方式"])
    chemistry.append(["任务编号", "元素", "最终值", "下限", "上限", "是否合格"])
    issues.append(["任务编号", "类型", "sheet", "行号", "字段", "code", "message", "suggestion"])
    rule_sheet.append(["参数键", "最终值"])

    rule_source = None
    for item in batch_result.get("results") or []:
        candidate = ((item.get("input") or {}).get("config") or {}).get("process_rules")
        if isinstance(candidate, dict):
            rule_source = candidate
            break
    if rule_source is None and isinstance(batch_result.get("processRules"), dict):
        rule_source = batch_result["processRules"]
    if isinstance(rule_source, dict):
        for label, key, _ in rule_template_rows():
            rule_sheet.append([key, rule_template_value(rule_source, key)])

    for item in batch_result.get("results") or []:
        task = item["input"]
        if item["status"] == "ok":
            mode = item["result"]["modes"]["milp"]
            manual_aluminum = normalized_manual_aluminum(task.get("manualAluminum"))
            heat_weight = float((task.get("config") or {}).get("heat_weight_t") or item["result"].get("heatWeightT") or 0)
            auto_cost = float(mode.get("costPerTon") or 0)
            auto_kg = float(mode.get("totalKgPerTon") or 0)
            manual_kg = manual_aluminum["kgPerTon"]
            manual_cost = manual_aluminum_cost_per_ton(manual_aluminum)
            total_cost = auto_cost
            total_kg = auto_kg
            summary.append([task["taskId"], task["grade"], task["thicknessMm"], "成功", auto_cost, manual_cost, total_cost, auto_kg, manual_kg, total_kg, total_cost * heat_weight, ""])
            feed_modes = {alloy["name"]: alloy.get("feed_mode") for alloy in task.get("config", {}).get("alloys", [])}
            routed_alloys = [alloy for alloy in mode["alloys"] if float(alloy.get("kgPerTon") or 0) > 1e-8]
            routed_alloys.sort(key=lambda alloy: (-float(alloy.get("costPerTon") or 0), str(alloy.get("name") or "")))
            for route_index, alloy in enumerate(routed_alloys, 1):
                feed_mode = alloy.get("feedMode") or alloy.get("feed_mode") or feed_modes.get(alloy["name"]) or ""
                details.append([task["taskId"], route_index, alloy["name"], alloy["kgPerTon"], alloy["heatKg"], alloy["bags"], alloy["costPerTon"], feed_mode])
            if manual_kg > 1e-8:
                details.append(
                    [
                        task["taskId"],
                        len(routed_alloys) + 1,
                        manual_aluminum["materialName"] or "铝块",
                        manual_kg,
                        manual_kg * heat_weight,
                        None,
                        manual_cost,
                        "手工录入",
                    ]
                )
            for check in mode["chemistryChecks"]:
                chemistry.append([task["taskId"], check["element"], check["value"], check["min"], check["max"], "是" if check["ok"] else "否"])
            for warning in item["result"].get("warnings") or []:
                issues.append([task["taskId"], "warning", "", "", "", "OPTIMIZER_WARNING", warning, "人工确认后执行。"])
        else:
            message = "；".join(error["message"] for error in item.get("errors") or [])
            summary.append([task["taskId"], task["grade"], task["thicknessMm"], "失败", None, None, None, None, None, None, None, message])
            for error in item.get("errors") or []:
                issues.append([task["taskId"], "error", error.get("sheet"), error.get("row"), error.get("field"), error.get("code"), error.get("message"), error.get("suggestion")])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            width = min(36, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_task_rows(rows: list[dict], errors: list[dict]) -> list[dict]:
    tasks = []
    seen: set[str] = set()
    for row in rows:
        task_id = required_text(row, "任务编号", "01_批量任务", errors)
        if task_id in seen:
            errors.append(make_issue("01_批量任务", row["_row"], "任务编号", "DUPLICATE_TASK_ID", f"任务编号重复：{task_id}", "请保证任务编号唯一。"))
        seen.add(task_id)
        tasks.append(
            {
                "row": row["_row"],
                "taskId": task_id,
                "grade": required_text(row, "适用牌号", "01_批量任务", errors),
                "thicknessMm": required_number(row, "厚度mm", "01_批量任务", errors, minimum=0),
                "heatWeightT": required_number(row, "炉重t", "01_批量任务", errors, minimum=0),
                "priceScheme": required_text(row, "价格方案", "01_批量任务", errors),
                "steelmakingGrade": required_text(row, "炼钢牌号", "01_批量任务", errors),
                "manualAluminumKgPerT": optional_number(row.get(MANUAL_ALUMINUM_FIELD), "01_批量任务", row["_row"], MANUAL_ALUMINUM_FIELD, errors, minimum=0, maximum=50, allow_zero=True) or 0,
                "notes": optional_text(row.get("备注")) or "",
            }
        )
    return tasks


def parse_target_rows(rows: list[dict], errors: list[dict], process_rules: dict | None = None) -> list[dict]:
    targets = []
    for row in rows:
        target_spec: dict[str, dict[str, float | str | None]] = {}
        single_target_elements: set[str] = set()
        for header, value in row.items():
            parsed = element_header(header, ("目标",))
            if parsed is None:
                continue
            element, _ = parsed
            if element in TARGET_IGNORED_ELEMENTS:
                continue
            number = optional_number(value, "02_目标成分上下限", row["_row"], header, errors, minimum=0, maximum=10)
            if number is None:
                continue
            target_spec[element] = {"mode": "single", "value": number}
            single_target_elements.add(element)
        for header, value in row.items():
            parsed = element_header(header, ("下限", "上限"))
            if parsed is None:
                continue
            element, bound_name = parsed
            if element in TARGET_IGNORED_ELEMENTS:
                continue
            if element in single_target_elements:
                if optional_text(value) is not None:
                    errors.append(
                        make_issue(
                            "02_目标成分上下限",
                            row["_row"],
                            element,
                            "TARGET_COLUMN_CONFLICT",
                            f"{element} 不能同时填写目标列和上下限列。",
                            "同一元素请二选一：使用单值目标，或使用旧上下限列。",
                        )
                    )
                continue
            number = optional_number(value, "02_目标成分上下限", row["_row"], header, errors, minimum=0, maximum=10)
            if number is None:
                continue
            key = "min" if bound_name == "下限" else "max"
            spec = target_spec.setdefault(element, {"mode": "range", "min": None, "max": None})
            spec["mode"] = "range"
            spec[key] = number
        for element, spec in target_spec.items():
            if spec.get("mode") != "range":
                continue
            if spec.get("min") not in (None, 0) and spec.get("max") not in (None, 0) and float(spec["min"]) > float(spec["max"]):
                errors.append(make_issue("02_目标成分上下限", row["_row"], element, "TARGET_RANGE_CROSSED", f"{element} 下限大于上限。", "请修正目标成分上下限。"))
        preview_config = {
            "target_spec": deepcopy(target_spec),
            "process_rules": deepcopy(process_rules) if isinstance(process_rules, dict) else default_process_rules(),
            "control_targets": {"enabled": False, "margin": 0, "elements": {}},
            "safety_margins": {},
            "alloys": [],
        }
        targets.append({**business_key(row, "02_目标成分上下限", errors), "targetSpec": target_spec, "target": compile_rule_view(preview_config).as_legacy_target()})
    return targets


def validate_unique_steelmaking_grade(records: list[dict], sheet: str, errors: list[dict]) -> None:
    seen: dict[str, int] = {}
    for record in records:
        steelmaking_grade = record.get("steelmakingGrade")
        if not steelmaking_grade:
            continue
        if steelmaking_grade in seen:
            errors.append(
                make_issue(
                    sheet,
                    record.get("row"),
                    "炼钢牌号",
                    "DUPLICATE_STEELMAKING_GRADE",
                    f"炼钢牌号重复：{steelmaking_grade}。",
                    "炼钢牌号在该 sheet 内必须唯一，请删除或合并重复记录。",
                )
            )
            continue
        seen[steelmaking_grade] = record.get("row")


def target_bounds_from_single_value(element: str, value: float, process_rules: dict | None = None) -> dict[str, float]:
    """把外部单值目标转换成优化器需要的上下限。"""

    preview = compile_rule_view(
        {
            "target_spec": {element: {"mode": "single", "value": normalized_float(value)}},
            "process_rules": deepcopy(process_rules) if isinstance(process_rules, dict) else default_process_rules(),
            "control_targets": {"enabled": False, "margin": 0, "elements": {}},
            "safety_margins": {},
            "alloys": [],
        }
    )
    return preview.as_legacy_target().get(element, {})


def single_target_values_to_bounds(values_by_element: dict[str, float], process_rules: dict | None = None) -> dict[str, dict[str, float]]:
    """批量转换外部单值目标，供模板样例和兼容解析复用。"""

    return {element: target_bounds_from_single_value(element, value, process_rules) for element, value in values_by_element.items() if element not in TARGET_IGNORED_ELEMENTS}


def normalized_float(value: float) -> float:
    """避免浮点噪声进入业务配置。"""

    return round(float(value), 10)


def parse_endpoint_rows(rows: list[dict], errors: list[dict]) -> list[dict]:
    endpoints = []
    for row in rows:
        residual: dict[str, float] = {}
        recovery: dict[str, float] = {}
        for header, value in row.items():
            residual_header = element_header(header, ("终点",))
            recovery_header = element_header(header, ("回收率",))
            if residual_header is not None:
                element, _ = residual_header
                number = optional_number(value, "03_转炉终点与回收率", row["_row"], header, errors, minimum=0, maximum=10)
                if number is not None:
                    residual[element] = number
            if recovery_header is not None:
                element, _ = recovery_header
                if element in FIXED_RECOVERY_RATES:
                    continue
                number = optional_number(value, "03_转炉终点与回收率", row["_row"], header, errors, minimum=0, maximum=1.2)
                if number is not None:
                    recovery[element] = number
        apply_field_confirmed_recovery_overrides(row, recovery)
        endpoints.append({**business_key(row, "03_转炉终点与回收率", errors), "residual": residual, "recoveryRates": recovery})
    return endpoints


def apply_field_confirmed_recovery_overrides(row: dict, recovery: dict[str, float]) -> None:
    """修正已由现场确认的数据录入错误。"""

    steelmaking_grade = (optional_text(row.get("炼钢牌号")) or "").upper()
    for (grade, element), value in FIELD_CONFIRMED_RECOVERY_OVERRIDES.items():
        if steelmaking_grade == grade and float(recovery.get(element) or 0) == 0:
            recovery[element] = value


def parse_alloy_rows(rows: list[dict], errors: list[dict], warnings: list[dict]) -> list[dict]:
    alloys = []
    for row in rows:
        composition = {}
        for header, value in row.items():
            header_name = normalize_header(header)
            if not header_name or header_name.startswith("_") or header_name in ALLOY_METADATA_HEADERS:
                continue
            if header_name in IGNORED_ELEMENTS:
                continue
            number = optional_number(value, "04_合金成分库", row["_row"], header_name, errors, minimum=0, maximum=100)
            if number is not None:
                composition[header_name] = number
        enabled = parse_enabled(row.get("启用"), "04_合金成分库", row["_row"], errors)
        if enabled:
            for impurity in ("P", "S"):
                if impurity not in composition:
                    warnings.append(make_issue("04_合金成分库", row["_row"], impurity, "MISSING_IMPURITY_COMPOSITION", f"启用合金缺少 {impurity} 杂质含量。", "正式使用前建议补充化验值。"))
        bag_size = optional_number(row.get("袋重kg"), "04_合金成分库", row["_row"], "袋重kg", errors, minimum=0, allow_zero=True)
        feed_mode = parse_feed_mode(row, bag_size, errors, warnings)
        if feed_mode == "连续":
            if bag_size and bag_size > 0:
                errors.append(make_issue("04_合金成分库", row["_row"], "袋重kg", "CONTINUOUS_BAG_SIZE_MUST_BE_ZERO", "连续投料的袋重kg必须留空或填 0。", "如果需要按袋投料，请把投料方式改为 整袋；否则袋重kg填 0 或留空。"))
            bag_size = 0
        elif feed_mode == "整袋" and (bag_size is None or bag_size <= 0):
            errors.append(make_issue("04_合金成分库", row["_row"], "袋重kg", "WHOLE_BAG_SIZE_REQUIRED", "整袋投料的袋重kg必须大于 0。", "请填写实际袋重kg，或把投料方式改为 连续。"))

        alloys.append(
            {
                "row": row["_row"],
                "name": required_text(row, "合金名称", "04_合金成分库", errors),
                "priceMaterialName": required_text(row, "价格物料名", "04_合金成分库", errors),
                "enabled": enabled,
                "feedMode": feed_mode,
                "bagSizeKg": float(bag_size or 0),
                "maxAddKgPerT": required_number(row, "最大投加kg每t", "04_合金成分库", errors, minimum=0),
                "composition": composition,
                "notes": optional_text(row.get("备注")) or "",
            }
        )
    return alloys


def parse_feed_mode(row: dict, bag_size: float | None, errors: list[dict], warnings: list[dict]) -> str:
    """解析投料方式；兼容缺少投料方式的一版旧模板。"""

    feed_mode = optional_text(row.get("投料方式"))
    if "投加顺序" in row:
        warnings.append(make_issue("04_合金成分库", row["_row"], "投加顺序", "LEGACY_ADDITION_SEQUENCE_IGNORED", "旧模板中的投加顺序已忽略。", "新模板不需要填写投加顺序；路线序号只用于结果导出排序。"))
    if feed_mode is None:
        if "投加顺序" not in row:
            errors.append(make_issue("04_合金成分库", row["_row"], "投料方式", "MISSING_FEED_MODE", "投料方式不能为空。", "请填写 连续 或 整袋。"))
            return ""
        inferred = "整袋" if (bag_size is not None and bag_size > 0) else "连续"
        warnings.append(make_issue("04_合金成分库", row["_row"], "投料方式", "LEGACY_FEED_MODE_INFERRED", f"旧模板缺少投料方式，已按袋重kg推断为 {inferred}。", "建议下载新模板并明确填写 连续 或 整袋。"))
        return inferred
    if feed_mode not in FEED_MODES:
        errors.append(make_issue("04_合金成分库", row["_row"], "投料方式", "INVALID_FEED_MODE", f"投料方式只能填写 连续 或 整袋，当前为：{feed_mode}", "请填写 连续 或 整袋。"))
        return ""
    return feed_mode


def parse_price_rows(rows: list[dict], errors: list[dict]) -> list[dict]:
    prices = []
    seen = set()
    for row in rows:
        scheme = required_text(row, "价格方案", "05_价格表", errors)
        material = required_text(row, "物料名称", "05_价格表", errors)
        key = (scheme, material)
        if key in seen:
            errors.append(make_issue("05_价格表", row["_row"], "物料名称", "DUPLICATE_PRICE", f"价格重复：{scheme}/{material}", "同一价格方案下每个物料只能有一条价格。"))
        seen.add(key)
        prices.append(
            {
                "row": row["_row"],
                "priceScheme": scheme,
                "materialName": material,
                "priceDate": optional_text(row.get("价格日期")) or "",
                "pricePerTon": required_number(row, "价格元每吨", "05_价格表", errors, minimum=0),
            }
        )
    return prices


def match_business_row(records: list[dict], task: dict, sheet: str, errors: list[dict]) -> dict | None:
    steelmaking_grade = task.get("steelmakingGrade")
    matches = [item for item in records if item.get("steelmakingGrade") == steelmaking_grade]
    if not matches:
        errors.append(
            make_issue(
                "01_批量任务",
                task["row"],
                "炼钢牌号",
                "MATCH_NOT_FOUND",
                f"{sheet} 中找不到炼钢牌号 {steelmaking_grade} 对应记录。",
                "请补充该炼钢牌号记录，或修正任务里的炼钢牌号。",
            )
        )
        return None
    if len(matches) > 1:
        errors.append(
            make_issue(
                "01_批量任务",
                task["row"],
                "炼钢牌号",
                "DUPLICATE_STEELMAKING_GRADE",
                f"{sheet} 中炼钢牌号 {steelmaking_grade} 匹配到多条记录。",
                "炼钢牌号必须唯一，请删除或合并重复业务记录。",
            )
        )
        return None
    return matches[0]


def build_residual(target: dict, endpoint_residual: dict, alloys: list[dict]) -> dict[str, float]:
    elements = ordered_elements(target, alloys)
    residual = {element: 0.0 for element in elements}
    for element in ("C", "Mn"):
        if element in residual:
            residual[element] = float(endpoint_residual.get(element, 0) or 0)
    return residual


def build_recovery_rates(target: dict, endpoint_recovery: dict, endpoint_row: int | None, alloys: list[dict], errors: list[dict]) -> dict[str, float]:
    rates = dict(endpoint_recovery)
    rates.update(FIXED_RECOVERY_RATES)
    ignored_recovery_elements = IGNORED_ELEMENTS | TARGET_IGNORED_ELEMENTS
    for element in ordered_elements(target, alloys):
        if element in ignored_recovery_elements:
            continue
        if element not in rates:
            errors.append(make_issue("03_转炉终点与回收率", endpoint_row, f"{element}回收率", "RECOVERY_NOT_FOUND", f"缺少 {element} 回收率。", "请在转炉终点与回收率表补充该元素回收率。"))
    return {element: float(rates[element]) for element in rates if element not in ignored_recovery_elements}


def ordered_elements(target: dict, alloys: list[dict]) -> list[str]:
    seen: list[str] = []
    for source in [target, *(alloy.get("composition") or {} for alloy in alloys)]:
        for element in source:
            if element in IGNORED_ELEMENTS:
                continue
            if element not in seen:
                seen.append(element)
    return seen


def business_key(row: dict, sheet: str, errors: list[dict]) -> dict:
    steelmaking_grade = required_text(row, "炼钢牌号", sheet, errors)
    return {
        "row": row["_row"],
        "grade": optional_text(row.get("适用牌号")) or "",
        "minThicknessMm": 0.0,
        "maxThicknessMm": 0.0,
        "steelmakingGrade": steelmaking_grade,
    }


def optimizer_error_to_issue(task: dict, exc: OptimizerError) -> dict:
    details = exc.details or [str(exc)]
    field = None
    for element in ERROR_FIELD_ELEMENTS:
        if any(element in detail for detail in details):
            field = element
            break
    return make_issue("01_批量任务", task.get("row"), field, "OPTIMIZER_VALIDATION_ERROR", "；".join(details), "请按错误内容修正模板输入。")


def required_text(row: dict, field: str, sheet: str, errors: list[dict]) -> str:
    value = optional_text(row.get(field))
    if not value:
        errors.append(make_issue(sheet, row.get("_row"), field, "REQUIRED_FIELD_EMPTY", f"{field} 不能为空。", "请填写该字段。"))
        return ""
    return value


def optional_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def required_number(row: dict, field: str, sheet: str, errors: list[dict], minimum: float | None = None, maximum: float | None = None, allow_zero: bool = False) -> float:
    number = optional_number(row.get(field), sheet, row.get("_row"), field, errors, minimum=minimum, maximum=maximum, allow_zero=allow_zero)
    if number is None:
        errors.append(make_issue(sheet, row.get("_row"), field, "REQUIRED_FIELD_EMPTY", f"{field} 不能为空。", "请填写有效数字。"))
        return 0.0
    return number


def optional_number(value, sheet: str, row: int | None, field: str, errors: list[dict], minimum: float | None = None, maximum: float | None = None, allow_zero: bool = False) -> float | None:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, bool):
        errors.append(make_issue(sheet, row, field, "INVALID_NUMBER", f"{field} 不是有效数字。", "请填写数字，不要填写布尔值或文本。"))
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        errors.append(make_issue(sheet, row, field, "INVALID_NUMBER", f"{field} 不是有效数字：{value}", "请删除中文单位、空格或错误字符。"))
        return None
    if number != number or number in (float("inf"), float("-inf")):
        errors.append(make_issue(sheet, row, field, "INVALID_NUMBER", f"{field} 不是有限数字。", "请填写普通数字。"))
        return None
    if minimum is not None and number < minimum:
        errors.append(make_issue(sheet, row, field, "NUMBER_OUT_OF_RANGE", f"{field} 必须大于等于 {minimum}。", "请修正数值范围。"))
    if minimum == 0 and number == 0 and not allow_zero and field in {"厚度mm", "炉重t", "最大厚度mm", "最大投加kg每t", "投加顺序", "价格元每吨"}:
        errors.append(make_issue(sheet, row, field, "NUMBER_OUT_OF_RANGE", f"{field} 必须大于 0。", "请修正数值范围。"))
    if maximum is not None and number > maximum:
        errors.append(make_issue(sheet, row, field, "NUMBER_OUT_OF_RANGE", f"{field} 不能大于 {maximum}。", "请确认单位是否填错。"))
    return number


def parse_enabled(value, sheet: str, row: int, errors: list[dict]) -> bool:
    text = optional_text(value)
    if text is None:
        errors.append(make_issue(sheet, row, "启用", "REQUIRED_FIELD_EMPTY", "启用不能为空。", "请填写 是/否。"))
        return False
    normalized = text.lower()
    if normalized in {"是", "启用", "true", "1", "yes", "y"}:
        return True
    if normalized in {"否", "禁用", "false", "0", "no", "n"}:
        return False
    errors.append(make_issue(sheet, row, "启用", "INVALID_BOOLEAN", f"启用字段无法识别：{text}", "请填写 是 或 否。"))
    return False


def parse_bool_flag(value, sheet: str, row: int, field: str, errors: list[dict]) -> bool | None:
    text = optional_text(value)
    if text is None:
        return None
    normalized = text.lower()
    if normalized in {"是", "启用", "true", "1", "yes", "y"}:
        return True
    if normalized in {"否", "禁用", "false", "0", "no", "n"}:
        return False
    errors.append(make_issue(sheet, row, field, "INVALID_BOOLEAN", f"{field} 字段无法识别：{text}", "请填写 是 或 否。"))
    return None


def element_header(header: str, suffixes: tuple[str, ...]) -> tuple[str, str] | None:
    normalized = normalize_header(header)
    for suffix in suffixes:
        if normalized.endswith(suffix):
            element = normalized[: -len(suffix)].strip()
            if element in TEMPLATE_ELEMENTS:
                return element, suffix
    return None


def normalize_header(value) -> str:
    return "" if value is None else str(value).replace("\n", "").strip()


def is_empty_row(sheet, row_index: int) -> bool:
    return all(sheet.cell(row_index, column).value in (None, "") for column in range(1, sheet.max_column + 1))


def make_issue(sheet: str | None, row: int | None, field: str | None, code: str, message: str, suggestion: str) -> dict:
    return {"sheet": sheet, "row": row, "field": field, "code": code, "message": message, "suggestion": suggestion}
