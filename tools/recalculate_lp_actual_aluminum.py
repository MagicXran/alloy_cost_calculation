"""Recalculate alloy workbook rows with the current LP rules.

The output is intentionally one worksheet: row-level evidence, LP result,
comparison, alloy deltas, and explanation stay together for audit.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.batch_template import (
    FIELD_CONFIRMED_RECOVERY_OVERRIDES,
    FIXED_RECOVERY_RATES,
)
from app.core import (
    build_linear_model,
    chemistry_checks,
    chemistry_from_vector,
    cost_for,
    diagnose_infeasible,
    effective_bounds,
    evaluate_plan_against_rules,
)
from app.rules_engine import compile_rule_view
from app.solvers import get_solver


SOURCE_WORKBOOK = ROOT / "热卷成本效益测算20260613版（基础参数表）---发徐老师(3).xlsx"
DEFAULT_OUTPUT = ROOT / "outputs" / "lp_actual_aluminum" / "热卷成本效益测算20260613版_LP新算法_单源对比_20260617_成分结果.xlsx"
SOURCE_DATA_START_ROW = 5
ALUMINUM_SOURCE_COLUMN = "AH"

TARGET_COLUMNS = {
    "C": "I",
    "Si": "J",
    "Mn": "K",
    "P": "L",
    "S": "M",
    "V": "N",
    "Nb": "O",
    "Ti": "P",
    "Als": "Q",
    "Alt": "R",
    "Ca": "S",
    "Cr": "T",
    "Ni": "U",
    "Cu": "V",
    "Mo": "W",
    "B": "X",
    "Sb": "Y",
    "N": "Z",
}
RECOVERY_COLUMNS = {
    "Si": "K",
    "Mn": "L",
    "V": "M",
    "Nb": "N",
    "Ti": "O",
    "Cr": "P",
    "Ni": "Q",
    "Cu": "R",
    "Mo": "S",
    "B": "T",
    "Sb": "U",
    "P": "V",
    "S": "W",
    "C": "X",
}
ALLOY_COLUMNS = [
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AS",
    "AT",
    "AU",
]
IGNORED_TARGET_ELEMENTS = {"Ca", "N"}
SKIPPED_ALLOYS = {"稀土合金"}

# Composition values are the explicit constants used by the legacy workbook
# formula columns. Aluminum stays in the alloy list, but its LP variable is
# fixed to the same workbook's AH column because current process_rules mark it
# as manually maintained.
ALLOY_COMPOSITIONS = {
    "硅锰": {"C": 1.72, "Si": 17.69, "Mn": 65.66},
    "高碳锰铁": {"C": 6.69, "Mn": 74.60},
    "中碳锰铁": {"C": 1.50, "Mn": 78.54},
    "低碳锰铁": {"C": 0.64, "Mn": 81.19},
    "金属锰": {"Mn": 90.00},
    "硅铁合金": {"Si": 72.23},
    "铝块": {"Als": 99.00, "Alt": 99.00},
    "钒铁": {"V": 51.17},
    "铌铁": {"Nb": 65.04},
    "钛铁": {"Ti": 71.58},
    "高碳铬铁": {"C": 10.00, "Cr": 52.00},
    "低碳铬铁": {"C": 0.12, "Cr": 59.44},
    "镍板": {"Ni": 99.00},
    "铜板": {"Cu": 99.00},
    "钼铁": {"Mo": 57.70},
    "硼铁": {"B": 18.00},
    "锑锭": {"Sb": 99.50},
    "磷铁": {"P": 23.94},
    "硫铁": {"S": 29.00},
}


@dataclass(frozen=True)
class AluminumMatch:
    value: float | None
    source_row: int | None
    source_column: str
    method: str
    warning: str


def optional_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def rounded(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def format_delta(value: float | None, digits: int = 3) -> str:
    if value is None:
        return "空"
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.{digits}f}"


def load_process_rules() -> dict[str, Any]:
    config = json.loads((ROOT / "config.json").read_text(encoding="utf-8"))
    return config.get("process_rules") or {}


def process_rules_snapshot_text(rules: dict[str, Any]) -> str:
    thresholds = rules.get("trace_alloy_thresholds") or {}
    return (
        f"C余量={rules.get('carbon_target_margin')}; "
        f"禁硅={rules.get('disable_silicon_alloys_si_max')}; "
        f"低硅上限阈值={rules.get('single_target_si_upper_only_max')}; "
        f"Ti余量={rules.get('ti_safety_addition')}; "
        f"Ni/Cu/Mo/Sb/B={thresholds.get('Ni')}/{thresholds.get('Cu')}/{thresholds.get('Mo')}/{thresholds.get('Sb')}/{thresholds.get('B')}; "
        f"P/S禁投={rules.get('phosphorus_alloy_max')}/{rules.get('sulfur_alloy_max')}"
    )


def source_aluminum(cost_sheet, row: int) -> AluminumMatch:
    value = optional_float(cost_sheet[f"{ALUMINUM_SOURCE_COLUMN}{row}"].value)
    warning = "" if value is not None else f"1.合金成本!{ALUMINUM_SOURCE_COLUMN}{row} 铝块为空或非数字"
    return AluminumMatch(
        value=value,
        source_row=row,
        source_column=f"1.合金成本!{ALUMINUM_SOURCE_COLUMN}",
        method="同一工作簿AH列",
        warning=warning,
    )


def workbook_alloys(cost_sheet) -> list[dict[str, Any]]:
    alloys: list[dict[str, Any]] = []
    for column in ALLOY_COLUMNS:
        name = cost_sheet[f"{column}2"].value
        if not name or name in SKIPPED_ALLOYS:
            continue
        name = str(name)
        composition = ALLOY_COMPOSITIONS.get(name)
        if not composition:
            continue
        alloys.append(
            {
                "name": name,
                "source_column": column,
                "price_per_ton": optional_float(cost_sheet[f"{column}4"].value) or 0.0,
                "composition": composition,
                "enabled": True,
                "addition_sequence": 99,
                "bag_size_kg": 0,
                "max_add_kg_per_t": 200,
                "recovery_overrides": {},
                "notes": "",
            }
        )
    return alloys


def source_data_rows(cost_sheet) -> list[int]:
    return [
        row
        for row in range(SOURCE_DATA_START_ROW, cost_sheet.max_row + 1)
        if cost_sheet[f"F{row}"].value not in (None, "")
    ]


def source_range_text(rows: list[int]) -> str:
    if not rows:
        return "无数据行"
    return f"{min(rows)}:{max(rows)}"


def apply_field_recovery_overrides(grade: str, recovery: dict[str, float]) -> list[str]:
    applied: list[str] = []
    normalized_grade = grade.upper()
    for (target_grade, element), value in FIELD_CONFIRMED_RECOVERY_OVERRIDES.items():
        if normalized_grade == target_grade and float(recovery.get(element) or 0) == 0:
            recovery[element] = value
            applied.append(f"{element}=0->{value}")
    return applied


def build_row_config(
    cost_sheet,
    param_sheet,
    row: int,
    grade: str,
    alloys: list[dict[str, Any]],
    process_rules: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, float | None], list[str]]:
    target_spec: dict[str, dict[str, float | str | None]] = {}
    raw_targets: dict[str, float | None] = {}
    for element, column in TARGET_COLUMNS.items():
        if element in IGNORED_TARGET_ELEMENTS:
            continue
        value = optional_float(cost_sheet[f"{column}{row}"].value)
        raw_targets[element] = value
        if value is None:
            continue
        target_spec[element] = {"mode": "single", "value": value}

    recovery: dict[str, float] = {}
    for element, column in RECOVERY_COLUMNS.items():
        value = optional_float(param_sheet[f"{column}{row}"].value)
        if value is not None:
            recovery[element] = value
    recovery_overrides = apply_field_recovery_overrides(grade, recovery)
    recovery.update(FIXED_RECOVERY_RATES)

    residual = {
        "C": optional_float(param_sheet[f"I{row}"].value) or 0.0,
        "Mn": optional_float(param_sheet[f"J{row}"].value) or 0.0,
    }
    used_elements = set(target_spec)
    for alloy in alloys:
        used_elements.update(alloy["composition"])
    for element in used_elements:
        residual.setdefault(element, 0.0)
        recovery.setdefault(element, 1.0)

    preview_config = {
        "target_spec": target_spec,
        "process_rules": process_rules,
        "control_targets": {"enabled": False, "margin": 0, "elements": {}},
        "safety_margins": {element: {"low": 0, "high": 0} for element in target_spec},
        "alloys": alloys,
    }
    preview_view = compile_rule_view(preview_config)
    compiled_target = preview_view.as_legacy_target()

    config = {
        "heat_weight_t": 150,
        "steel_weight_kg": 1000,
        "ignored_elements": sorted(IGNORED_TARGET_ELEMENTS),
        "target_spec": preview_view.target_spec,
        "target": compiled_target,
        "raw_targets": raw_targets,
        "residual": residual,
        "recovery_rates": recovery,
        "safety_margins": {element: {"low": 0, "high": 0} for element in target_spec},
        "control_targets": {"enabled": False, "margin": 0, "elements": {}},
        "process_rules": process_rules,
        "alloys": alloys,
        "milp_settings": {"default_bag_size_kg": 25, "enable_bag_rounding": False},
        "temperature_drop": {"enabled": False},
    }
    return config, raw_targets, recovery_overrides


def constraint_text(config: dict[str, Any]) -> str:
    parts: list[str] = []
    for element in TARGET_COLUMNS:
        if element in IGNORED_TARGET_ELEMENTS:
            continue
        bounds = effective_bounds(config, element)
        if bounds["min"] is None and bounds["max"] is None:
            continue
        if bounds["min"] is not None and bounds["max"] is not None:
            parts.append(f"{element}[{bounds['min']:.6g},{bounds['max']:.6g}]")
        elif bounds["min"] is not None:
            parts.append(f"{element}>={bounds['min']:.6g}")
        else:
            parts.append(f"{element}<={bounds['max']:.6g}")
    return "; ".join(parts)


def bounds_snapshot(config: dict[str, Any]) -> dict[str, dict[str, float | None]]:
    return {
        element: effective_bounds(config, element)
        for element in TARGET_COLUMNS
        if element not in IGNORED_TARGET_ELEMENTS
    }


def chemistry_text(checks: list[dict[str, Any]]) -> str:
    shown = []
    for check in checks:
        minimum = check.get("min")
        maximum = check.get("max")
        if minimum is None and maximum is None:
            continue
        status = "OK" if check.get("ok") else "NG"
        shown.append(f"{check['element']}={check['value']:.6g}({status})")
    return "; ".join(shown)


def evaluate_excel_plan(
    config: dict[str, Any],
    alloys: list[dict[str, Any]],
    old_x: list[float | None],
    old_raw_x: list[Any],
) -> dict[str, Any]:
    """判断旧 Excel 投料在当前批准规则下是否可行。"""

    input_issues: list[str] = []
    normalized_x: list[float] = []
    for index, alloy in enumerate(alloys):
        raw = old_raw_x[index]
        value = old_x[index]
        if value is None:
            if raw not in (None, ""):
                input_issues.append(f"{alloy['name']} 原表投料非数字：{raw}")
            normalized_x.append(0.0)
            continue
        if value < -1e-8:
            input_issues.append(f"{alloy['name']} 原表投料为负：{value:.6g} kg/t")
        normalized_x.append(value)

    evaluation = evaluate_plan_against_rules(config, alloys, normalized_x, ignore_manual_aluminum=True)
    issues = [*input_issues, *evaluation["issues"]]
    if input_issues:
        status = "输入异常"
    elif evaluation["ok"]:
        status = "是"
    else:
        status = "否"
    return {
        "status": status,
        "checks": evaluation["checks"],
        "chemistry": evaluation["chemistry"],
        "issues": issues,
        "x": normalized_x,
    }


def top_alloy_changes(
    alloys: list[dict[str, Any]],
    old_x: list[float | None],
    new_x: list[float],
    limit: int = 5,
) -> list[str]:
    changes = []
    for index, alloy in enumerate(alloys):
        old_value = old_x[index]
        if old_value is None:
            if abs(new_x[index]) < 1e-6:
                continue
            changes.append((abs(new_x[index]), alloy["name"], None, new_x[index], None))
            continue
        delta = new_x[index] - old_value
        if abs(delta) < 1e-6:
            continue
        changes.append((abs(delta), alloy["name"], old_value, new_x[index], delta))
    changes.sort(reverse=True)
    formatted = []
    for _, name, old, new, delta in changes[:limit]:
        if old is None:
            formatted.append(f"{name}:Excel错误->{new:.3f}")
        else:
            formatted.append(f"{name}:{old:.3f}->{new:.3f}({format_delta(delta)})")
    return formatted


def source_audit_notes(
    cost_sheet,
    param_sheet,
    row: int,
    alloys: list[dict[str, Any]],
    old_x: list[float | None],
    old_total: float | None,
    old_cost: float | None,
) -> list[str]:
    notes: list[str] = []
    cost_grade = str(cost_sheet[f"F{row}"].value or "").strip()
    param_grade = str(param_sheet[f"F{row}"].value or "").strip()
    if cost_grade != param_grade:
        notes.append(f"炼钢牌号行对齐异常：1.合金成本!F{row}={cost_grade}，炼钢参数表!F{row}={param_grade}")

    for index, alloy in enumerate(alloys):
        value = old_x[index]
        raw = cost_sheet[f"{alloy['source_column']}{row}"].value
        if value is None and raw not in (None, ""):
            notes.append(f"{alloy['name']} 原表投料非数字：{raw}")
        elif value is not None and value < -1e-8:
            notes.append(f"{alloy['name']} 原表投料为负：{value:.6g} kg/t")

    if old_total is not None:
        total_from_alloys = sum(value or 0.0 for value in old_x)
        if abs(total_from_alloys - old_total) > 1e-6:
            notes.append(f"AV合金消耗与逐合金求和不一致：sum={total_from_alloys:.6f}，AV={old_total:.6f}")
    if old_cost is not None:
        cost_from_alloys = sum((value or 0.0) * float(alloy["price_per_ton"]) / 1000 for value, alloy in zip(old_x, alloys))
        if abs(cost_from_alloys - old_cost) > 1e-6:
            notes.append(f"AW合金成本与单价乘投料不一致：sumproduct={cost_from_alloys:.6f}，AW={old_cost:.6f}")

    return notes


def explain_row(
    status: str,
    grade: str,
    config: dict[str, Any],
    alloys: list[dict[str, Any]],
    old_x: list[float | None],
    new_x: list[float] | None,
    old_chemistry: dict[str, float] | None,
    excel_rule_status: str,
    excel_rule_issues: list[str],
    old_cost: float | None,
    old_cost_raw: Any,
    old_total_raw: Any,
    new_cost: float | None,
    aluminum: AluminumMatch,
    recovery_overrides: list[str],
    diagnostics: list[str],
    audit_notes: list[str],
) -> str:
    if status != "ok":
        return "不可行：" + "；".join(diagnostics)

    assert new_x is not None
    reasons: list[str] = []
    old_al = old_x[[alloy["name"] for alloy in alloys].index("铝块")]
    old_al = 0.0 if old_al is None else old_al
    actual_al = aluminum.value or 0.0
    if abs(actual_al - old_al) > 1e-6:
        reasons.append(
            f"铝块按同源AH维护：原Excel AH={old_al:.3f}kg/t，{aluminum.source_column}{aluminum.source_row}={actual_al:.3f}kg/t，差{format_delta(actual_al - old_al)}kg/t"
        )
    else:
        reasons.append(f"铝块按同源AH列固定计入：{actual_al:.3f}kg/t")

    if audit_notes:
        reasons.append("源表审计提示：" + "；".join(audit_notes[:3]))

    if excel_rule_status != "是":
        prefix = "原Excel在批准规则下输入异常：" if excel_rule_status == "输入异常" else "原Excel在批准规则下不可行："
        reasons.append(prefix + "；".join(excel_rule_issues[:3]))

    if recovery_overrides:
        reasons.append(f"现场确认回收率修正：{grade} " + "、".join(recovery_overrides))

    c_bounds = effective_bounds(config, "C")
    if c_bounds["max"] is not None:
        c_target = (config.get("raw_targets") or {}).get("C")
        carbon_margin = float(compile_rule_view(config).resolved_rules_config["carbon_target_margin"])
        prefix = f"C上限按目标{c_target:.3f}-{carbon_margin:.3f}={c_bounds['max']:.3f}"
        if old_chemistry and old_chemistry.get("C", 0.0) > c_bounds["max"] + 1e-6:
            reasons.append(prefix + f"，原Excel按当前成分口径估算C={old_chemistry['C']:.3f}，会被新约束压住")
        else:
            reasons.append(prefix)

    si_target = (config.get("raw_targets") or {}).get("Si")
    if si_target is not None and si_target <= float(config["process_rules"]["disable_silicon_alloys_si_max"]) + 1e-8:
        reasons.append(f"Si目标={si_target:.3f}<=禁硅阈值，硅锰/硅铁变量上限置0")

    changes = top_alloy_changes(alloys, old_x, new_x)
    if changes:
        reasons.append("主要投料变化：" + "；".join(changes))

    if old_cost is None:
        reasons.append(f"原Excel AV/AW结果无效：AV={old_total_raw}，AW={old_cost_raw}，本行新旧成本差不可比")
    elif new_cost is not None:
        delta = new_cost - old_cost
        direction = "下降" if delta < -1e-6 else "上升" if delta > 1e-6 else "基本持平"
        reasons.append(f"成本{direction}：新-旧={format_delta(delta)}元/t")

    if aluminum.warning:
        reasons.append("铝耗匹配提示：" + aluminum.warning)
    return "；".join(reasons)


def compute_rows(source_workbook_path: Path = SOURCE_WORKBOOK) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    process_rules = load_process_rules()
    source_workbook = openpyxl.load_workbook(source_workbook_path, data_only=True, read_only=False)
    cost_sheet = source_workbook["1.合金成本"]
    param_sheet = source_workbook["炼钢参数表"]
    rows_to_compute = source_data_rows(cost_sheet)
    alloys = workbook_alloys(cost_sheet)
    solver = get_solver("highs")

    rows: list[dict[str, Any]] = []
    details: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()

    for row in rows_to_compute:
        grade_cell = cost_sheet[f"F{row}"].value
        if not grade_cell:
            continue
        grade = str(grade_cell).strip()
        aluminum = source_aluminum(cost_sheet, row)
        config, raw_targets, recovery_overrides = build_row_config(cost_sheet, param_sheet, row, grade, alloys, process_rules)
        old_raw_x = [cost_sheet[f"{alloy['source_column']}{row}"].value for alloy in alloys]
        old_x = [optional_float(value) for value in old_raw_x]
        old_x_for_chemistry = [value if value is not None else 0.0 for value in old_x]
        old_cost_raw = cost_sheet[f"AW{row}"].value
        old_total_raw = cost_sheet[f"AV{row}"].value
        old_cost = optional_float(old_cost_raw)
        old_total = optional_float(old_total_raw)
        excel_result_status = "ok" if old_cost is not None and old_total is not None else f"Excel原表错误: AV={old_total_raw}, AW={old_cost_raw}"
        audit_notes = source_audit_notes(cost_sheet, param_sheet, row, alloys, old_x, old_total, old_cost)
        old_chemistry = chemistry_from_vector(config, alloys, old_x_for_chemistry)
        excel_rule_eval = evaluate_excel_plan(config, alloys, old_x, old_raw_x)
        constraints = constraint_text(config)

        diagnostics: list[str] = []
        status = "ok"
        new_x: list[float] | None = None
        new_auto_x: list[float] | None = None
        new_total: float | None = None
        new_cost: float | None = None
        new_auto_total: float | None = None
        new_auto_cost: float | None = None
        new_chemistry: dict[str, float] | None = None
        new_chemistry_checks: list[dict[str, Any]] = []

        raw_solution = solver.solve_lp(build_linear_model(config, alloys).as_dict())
        if raw_solution is None:
            status = "infeasible"
            diagnostics = diagnose_infeasible(config, alloys, solver)
            counters["infeasible"] += 1
        else:
            new_auto_x = [max(0.0, value or 0.0) for value in raw_solution.x]
            new_x = list(new_auto_x)
            for index, alloy in enumerate(alloys):
                if alloy["name"] == "铝块":
                    new_x[index] = aluminum.value or 0.0
                    new_auto_x[index] = 0.0
            new_total = sum(new_x)
            new_cost = cost_for(alloys, new_x)
            new_auto_total = sum(new_auto_x)
            new_auto_cost = cost_for(alloys, new_auto_x)
            new_chemistry = chemistry_from_vector(config, alloys, new_x)
            new_chemistry_checks = chemistry_checks(config, new_chemistry)
            counters["ok"] += 1

        if aluminum.value is None:
            counters["missing_aluminum"] += 1
        if aluminum.warning:
            counters["aluminum_warning"] += 1
        if recovery_overrides:
            counters["recovery_overrides"] += 1
        if audit_notes:
            counters["source_audit_warnings"] += 1
        if any("投料为负" in note for note in audit_notes):
            counters["negative_alloy_rows"] += 1
        if excel_rule_eval["status"] == "是":
            counters["excel_rule_ok"] += 1
        elif excel_rule_eval["status"] == "否":
            counters["excel_rule_ng"] += 1
        else:
            counters["excel_rule_input_error"] += 1

        explanation = explain_row(
            status,
            grade,
            config,
            alloys,
            old_x,
            new_x,
            old_chemistry,
            excel_rule_eval["status"],
            excel_rule_eval["issues"],
            old_cost,
            old_cost_raw,
            old_total_raw,
            new_cost,
            aluminum,
            recovery_overrides,
            diagnostics,
            audit_notes,
        )

        record = {
            "Excel行号": row,
            "序号": cost_sheet[f"A{row}"].value,
            "适用牌号": cost_sheet[f"B{row}"].value,
            "分类序号": cost_sheet[f"C{row}"].value,
            "最小厚度mm": cost_sheet[f"D{row}"].value,
            "最大厚度mm": cost_sheet[f"E{row}"].value,
            "炼钢牌号": grade,
            "钢种类型": cost_sheet[f"G{row}"].value,
            "液相线温度℃": cost_sheet[f"H{row}"].value,
            "目标": raw_targets,
            "终点C": config["residual"].get("C"),
            "终点Mn": config["residual"].get("Mn"),
            "终点成分": dict(config["residual"]),
            "成分边界": bounds_snapshot(config),
            "LP状态": status,
            "Excel结果状态": excel_result_status,
            "Excel是否满足批准规则": excel_rule_eval["status"],
            "Excel原方案成分校核": chemistry_text(excel_rule_eval["checks"]),
            "Excel不满足批准规则原因": "；".join(excel_rule_eval["issues"]),
            "源表审计状态": "OK" if not audit_notes else "WARN",
            "源表审计提示": "；".join(audit_notes),
            "Excel合金消耗kg/t": old_total,
            "新算法合金消耗kg/t": new_total,
            "新-Excel消耗kg/t": None if new_total is None or old_total is None else new_total - old_total,
            "Excel合金消耗原值": old_total_raw,
            "Excel合金成本元/t": old_cost,
            "新算法合金成本元/t": new_cost,
            "新-Excel成本元/t": None if new_cost is None or old_cost is None else new_cost - old_cost,
            "Excel合金成本原值": old_cost_raw,
            "Excel铝块kg/t": old_x[[alloy["name"] for alloy in alloys].index("铝块")],
            "同源铝耗kg/t": aluminum.value,
            "铝耗来源": "" if aluminum.source_row is None else f"{aluminum.source_column}{aluminum.source_row}",
            "铝耗口径": aluminum.method,
            "铝耗警告": aluminum.warning,
            "LP自动合金消耗kg/t(不含铝)": new_auto_total,
            "LP自动成本元/t(不含铝)": new_auto_cost,
            "LP目标约束": constraints,
            "新方案成分校核": chemistry_text(new_chemistry_checks),
            "原因": explanation,
            "诊断": "；".join(diagnostics),
            "old_x": old_x,
            "old_raw_x": old_raw_x,
            "new_x": new_x,
            "old_chemistry": old_chemistry,
            "new_chemistry": new_chemistry,
        }
        rows.append(record)

        if new_x is not None:
            for index, alloy in enumerate(alloys):
                old_value = old_x[index]
                delta = None if old_value is None else new_x[index] - old_value
                if old_value is None and abs(new_x[index]) < 1e-6:
                    continue
                if delta is not None and abs(delta) < 1e-6:
                    continue
                details.append(
                    {
                        "Excel行号": row,
                        "炼钢牌号": grade,
                        "合金": alloy["name"],
                        "Excel kg/t": old_value,
                        "新算法 kg/t": new_x[index],
                        "差异 kg/t": delta,
                        "Excel成本元/t": None if old_value is None else old_value * float(alloy["price_per_ton"]) / 1000,
                        "新算法成本元/t": new_x[index] * float(alloy["price_per_ton"]) / 1000,
                    }
                )

    ok_rows = [row for row in rows if row["LP状态"] == "ok" and row["Excel结果状态"] == "ok"]
    summary = {
        "total": len(rows),
        "ok": counters["ok"],
        "infeasible": counters["infeasible"],
        "excel_result_errors": sum(1 for row in rows if row["Excel结果状态"] != "ok"),
        "missing_aluminum": counters["missing_aluminum"],
        "aluminum_warnings": counters["aluminum_warning"],
        "recovery_overrides": counters["recovery_overrides"],
        "source_audit_warnings": counters["source_audit_warnings"],
        "negative_alloy_rows": counters["negative_alloy_rows"],
        "excel_rule_ok": counters["excel_rule_ok"],
        "excel_rule_ng": counters["excel_rule_ng"],
        "excel_rule_input_error": counters["excel_rule_input_error"],
        "old_cost_total": sum(row["Excel合金成本元/t"] or 0 for row in ok_rows),
        "new_cost_total": sum(row["新算法合金成本元/t"] or 0 for row in ok_rows),
        "old_kg_total": sum(row["Excel合金消耗kg/t"] or 0 for row in ok_rows),
        "new_kg_total": sum(row["新算法合金消耗kg/t"] or 0 for row in ok_rows),
    }
    summary["cost_delta_total"] = summary["new_cost_total"] - summary["old_cost_total"]
    summary["kg_delta_total"] = summary["new_kg_total"] - summary["old_kg_total"]
    return rows, details, {"summary": summary, "alloys": alloys, "source_workbook": source_workbook_path, "source_rows": rows_to_compute, "process_rules": process_rules}


def result_headers(alloys: list[dict[str, Any]]) -> list[str]:
    base_headers = [
        "Excel行号",
        "序号",
        "适用牌号",
        "分类序号",
        "最小厚度mm",
        "最大厚度mm",
        "炼钢牌号",
        "钢种类型",
        "液相线温度℃",
        *[f"{element}目标" for element in TARGET_COLUMNS if element not in IGNORED_TARGET_ELEMENTS],
        "终点C",
        "终点Mn",
        "LP状态",
        "Excel结果状态",
        "Excel是否满足批准规则",
        "Excel原方案成分校核",
        "Excel不满足批准规则原因",
        "源表审计状态",
        "源表审计提示",
        "Excel合金消耗kg/t",
        "Excel合金消耗原值",
        "新算法合金消耗kg/t",
        "新-Excel消耗kg/t",
        "Excel合金成本元/t",
        "Excel合金成本原值",
        "新算法合金成本元/t",
        "新-Excel成本元/t",
        "Excel铝块kg/t",
        "同源铝耗kg/t",
        "铝耗来源",
        "铝耗口径",
        "铝耗警告",
        "LP自动合金消耗kg/t(不含铝)",
        "LP自动成本元/t(不含铝)",
        "LP目标约束",
        "新方案成分校核",
        "原因",
        "诊断",
    ]
    alloy_headers: list[str] = []
    for alloy in alloys:
        name = alloy["name"]
        alloy_headers.extend([f"Excel_{name}kg/t", f"新算法_{name}kg/t", f"差异_{name}kg/t"])
    return base_headers + alloy_headers


def record_to_row(record: dict[str, Any], headers: list[str], alloys: list[dict[str, Any]]) -> list[Any]:
    values: dict[str, Any] = {header: record.get(header) for header in headers}
    for element in TARGET_COLUMNS:
        if element not in IGNORED_TARGET_ELEMENTS:
            values[f"{element}目标"] = (record.get("目标") or {}).get(element)
    for index, alloy in enumerate(alloys):
        name = alloy["name"]
        old_x = record.get("old_x") or []
        new_x = record.get("new_x") or []
        old_value = old_x[index] if index < len(old_x) else None
        new_value = new_x[index] if index < len(new_x) else None
        values[f"Excel_{name}kg/t"] = old_value
        values[f"新算法_{name}kg/t"] = new_value
        values[f"差异_{name}kg/t"] = None if new_value is None or old_value is None else new_value - old_value
    return [values.get(header) for header in headers]


def chemistry_status(value: float | None, bounds: dict[str, float | None] | None, *, missing_status: str = "空") -> str:
    if value is None:
        return missing_status
    bounds = bounds or {"min": None, "max": None}
    minimum = bounds.get("min")
    maximum = bounds.get("max")
    if minimum is None and maximum is None:
        return "无约束"
    if minimum is not None and value < minimum - 1e-8:
        return "NG"
    if maximum is not None and value > maximum + 1e-8:
        return "NG"
    return "OK"


def write_composition_sheet(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("成分结果")
    headers = [
        "Excel行号",
        "炼钢牌号",
        "元素",
        "目标值",
        "约束下限",
        "约束上限",
        "终点成分",
        "Excel方案成分",
        "新算法成分",
        "新-Excel成分",
        "Excel校核",
        "新算法校核",
        "LP状态",
        "Excel是否满足批准规则",
        "Excel不满足批准规则原因",
        "LP目标约束",
        "序号",
        "适用牌号",
        "钢种类型",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14532D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    output_row = 2
    for record in rows:
        targets = record.get("目标") or {}
        bounds_by_element = record.get("成分边界") or {}
        residual = record.get("终点成分") or {}
        old_chemistry = record.get("old_chemistry") or {}
        new_chemistry = record.get("new_chemistry") or {}
        for element in TARGET_COLUMNS:
            if element in IGNORED_TARGET_ELEMENTS:
                continue
            bounds = bounds_by_element.get(element) or {"min": None, "max": None}
            old_value = old_chemistry.get(element)
            new_value = new_chemistry.get(element)
            delta = None if old_value is None or new_value is None else new_value - old_value
            row_values = [
                record.get("Excel行号"),
                record.get("炼钢牌号"),
                element,
                targets.get(element),
                bounds.get("min"),
                bounds.get("max"),
                residual.get(element),
                old_value,
                new_value,
                delta,
                chemistry_status(old_value, bounds),
                chemistry_status(new_value, bounds, missing_status="无解" if record.get("LP状态") != "ok" else "空"),
                record.get("LP状态"),
                record.get("Excel是否满足批准规则"),
                record.get("Excel不满足批准规则原因"),
                record.get("LP目标约束"),
                record.get("序号"),
                record.get("适用牌号"),
                record.get("钢种类型"),
            ]
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(output_row, column_index, rounded(value) if isinstance(value, float) else value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column_index in (11, 12):
                    if value == "NG":
                        cell.fill = PatternFill("solid", fgColor="FCA5A5")
                    elif value == "OK":
                        cell.fill = PatternFill("solid", fgColor="DCFCE7")
                if column_index == 10 and isinstance(value, (int, float)):
                    if value < -1e-8:
                        cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    elif value > 1e-8:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
            output_row += 1

    if output_row > 2:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{output_row - 1}"
    sheet.freeze_panes = "A2"
    widths = {
        "A": 10,
        "B": 16,
        "C": 8,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 13,
        "I": 13,
        "J": 13,
        "K": 12,
        "L": 12,
        "M": 10,
        "N": 18,
        "O": 48,
        "P": 44,
        "Q": 9,
        "R": 14,
        "S": 14,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for row in sheet.iter_rows(min_row=2, max_row=output_row - 1, min_col=4, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
    sheet.row_dimensions[1].height = 32


def write_workbook(rows: list[dict[str, Any]], details: list[dict[str, Any]], meta: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "LP新算法铝耗对比"

    summary = meta["summary"]
    alloys = meta["alloys"]
    source_workbook = Path(meta["source_workbook"])
    source_rows = meta.get("source_rows") or []
    source_row_range = source_range_text(source_rows)
    headers = result_headers(alloys)
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, len(headers)))
    sheet["A1"] = f"{source_workbook.name} - LP新算法单源对比"
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")

    summary_rows = [
        (
            "源文件",
            f"{source_workbook.name}!1.合金成本 第{source_row_range}行, AB4:AU4单价, AH列铝块；炼钢参数表 第{source_row_range}行",
            "本次不读取外部铝耗表或合金价格表，所有输入来自同一个 workbook。",
        ),
        ("LP口径", "当前 app.core + rules_engine + target_spec", "铝块 manual_aluminum，不参与LP自动优化；新方案固定使用同源 AH 铝块值计入总耗和成本。"),
        (
            "行数",
            summary["total"],
            f"数据行 {source_row_range}；LP可行 {summary['ok']}；不可行 {summary['infeasible']}；Excel原结果错误 {summary['excel_result_errors']}；批准规则下 Excel 可行 {summary['excel_rule_ok']}、不可行 {summary['excel_rule_ng']}、输入异常 {summary['excel_rule_input_error']}；AH铝耗缺失 {summary['missing_aluminum']}。",
        ),
        ("规则参数", process_rules_snapshot_text(meta["process_rules"]), "本次回算先读取 config.json 默认规则，再按该快照统一编译目标语义、禁投逻辑和元素边界。"),
        (
            "成本累计",
            round(summary["new_cost_total"], 6),
            f"仅统计Excel原结果有效行：原Excel {summary['old_cost_total']:.6f}；新-旧 {summary['cost_delta_total']:.6f} 元/t累计。",
        ),
        (
            "消耗累计",
            round(summary["new_kg_total"], 6),
            f"仅统计Excel原结果有效行：原Excel {summary['old_kg_total']:.6f}；新-旧 {summary['kg_delta_total']:.6f} kg/t累计。",
        ),
        ("源表审计", summary["source_audit_warnings"], f"有审计提示的行数；其中原表投料为负 {summary['negative_alloy_rows']} 行；AW/AV 会按第4行单价和逐合金投料复算校验。"),
        ("特殊修正", summary["recovery_overrides"], "现场确认回收率修正计数；正确版 26MnB5 的 Si 回收率已为 0.8 时不会触发覆盖。"),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "本表由 tools/recalculate_lp_actual_aluminum.py 生成。"),
    ]
    for row_index, row_values in enumerate(summary_rows, start=2):
        label, value, note = row_values
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 9, note)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        sheet.merge_cells(start_row=row_index, start_column=9, end_row=row_index, end_column=17)
        for column_index in (1, 2, 9):
            cell = sheet.cell(row_index, column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 1).font = Font(bold=True)

    header_row = len(summary_rows) + 3
    group_row = header_row - 1
    sheet.cell(group_row, 1, "基本信息/目标/旧结果/铝耗/新LP/差异/原因/逐合金明细")
    sheet.cell(group_row, 1).font = Font(bold=True, color="FFFFFF")
    sheet.cell(group_row, 1).fill = PatternFill("solid", fgColor="5B9BD5")
    sheet.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=len(headers))

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, record in enumerate(rows, start=1):
        row_index = header_row + offset
        for column_index, value in enumerate(record_to_row(record, headers, alloys), start=1):
            cell = sheet.cell(row_index, column_index, rounded(value) if isinstance(value, float) else value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[column_index - 1] == "LP状态" and value != "ok":
                cell.fill = PatternFill("solid", fgColor="FCA5A5")
            if headers[column_index - 1].startswith("新-Excel") and isinstance(value, (int, float)):
                if value < -1e-6:
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                elif value > 1e-6:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")

    sheet.freeze_panes = f"J{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + len(rows)}"

    widths = {
        "A": 10,
        "B": 9,
        "C": 12,
        "D": 10,
        "E": 12,
        "F": 12,
        "G": 16,
        "H": 34,
        "I": 13,
    }
    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        if letter in widths:
            width = widths[letter]
        elif header.endswith("目标"):
            width = 10
        elif "原因" in header or "约束" in header or "校核" in header or "警告" in header or "诊断" in header or "审计提示" in header:
            width = 45
        elif header.startswith(("Excel_", "新算法_", "差异_")):
            width = 13
        else:
            width = 16
        sheet.column_dimensions[letter].width = width

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + len(rows)):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
    for row_index in range(2, header_row - 1):
        sheet.row_dimensions[row_index].height = 32
    sheet.row_dimensions[header_row].height = 42

    write_composition_sheet(workbook, rows)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Recalculate LP comparison from the maintained single source workbook.")
    parser.add_argument("--source", type=Path, default=SOURCE_WORKBOOK, help="Authoritative source .xlsx path.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output .xlsx path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows, details, meta = compute_rows(args.source)
    write_workbook(rows, details, meta, args.output)
    summary = meta["summary"]
    print(
        json.dumps(
            {
                "output": str(args.output),
                "rows": summary["total"],
                "ok": summary["ok"],
                "infeasible": summary["infeasible"],
                "excel_result_errors": summary["excel_result_errors"],
                "excel_rule_ok": summary["excel_rule_ok"],
                "excel_rule_ng": summary["excel_rule_ng"],
                "excel_rule_input_error": summary["excel_rule_input_error"],
                "missing_aluminum": summary["missing_aluminum"],
                "source_audit_warnings": summary["source_audit_warnings"],
                "negative_alloy_rows": summary["negative_alloy_rows"],
                "cost_delta_total": round(summary["cost_delta_total"], 6),
                "kg_delta_total": round(summary["kg_delta_total"], 6),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
